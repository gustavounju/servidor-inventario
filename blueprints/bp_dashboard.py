from flask import Blueprint, render_template, request, redirect, url_for, abort, send_file, flash, session, jsonify
from datetime import datetime as dt
from database.db_core import get_db_connection
import socket
import os
from utils.constants import FUERO_COLORS, list_fuero_mapping_rows
import datetime
from io import BytesIO
from openpyxl import Workbook
from services.audit import log_audit_event
from services.dashboard_overview import load_dashboard_overview
from services.pc_actions import decommission_pc_service, reactivate_pc_service, delete_permanent_pc_service, update_pc_infrastructure_service
from services.pc_details_service import get_pc_detail_context
from services.asset_validation import is_ignored_storage_component
from services.fuero_service import get_fuero_summary_data, get_fuero_detail_data, recalculate_all_pc_fueros
from utils.auth import is_authenticated, login_required, permission_required, current_technician_identity, has_permission, current_username
from utils.component_status import deployed_component_state

bp_dashboard = Blueprint('dashboard', __name__)

@bp_dashboard.app_template_filter('normalize_ram_spec')
def normalize_ram_spec_filter(spec, fallback_gb=None, processor=None):
    from services.reporting import normalize_ram_spec
    return normalize_ram_spec(spec, fallback_gb, processor)



def _normalize_component_type(value):
    return (value or "").strip().upper()


def _acta_component_bucket(component_type):
    ctype = _normalize_component_type(component_type)
    if "GABINETE" in ctype or "CHASIS" in ctype or ctype == "CPU":
        return "gabinetes"
    if "FUENTE" in ctype or "PSU" in ctype or "POWER" in ctype:
        return "fuentes"
    if "PROCESADOR" in ctype or "MICRO" in ctype:
        return "procesadores"
    if "RAM" in ctype or "MEMORIA" in ctype:
        return "memorias"
    if "MOTHERBOARD" in ctype or "PLACA MADRE" in ctype:
        return "placas_madre"
    if "DISCO" in ctype or "ALMACENAMIENTO" in ctype or "SSD" in ctype or "HDD" in ctype:
        return "discos"
    if "MONITOR" in ctype or "PANTALLA" in ctype:
        return "monitores"
    return "otros"


def build_acta_component_groups(components, monitors_detail, hardware_components=None):
    groups = {
        "gabinetes": [],
        "fuentes": [],
        "procesadores": [],
        "memorias": [],
        "placas_madre": [],
        "discos": [],
        "otros": [],
        "monitores": [],
    }

    for raw_component in components or []:
        component = dict(raw_component)
        bucket = _acta_component_bucket(component.get("component_type"))
        if bucket == "monitores":
            continue
        groups[bucket].append(component)

    for raw_monitor in monitors_detail or []:
        groups["monitores"].append(dict(raw_monitor))

    # Los discos WMI forman parte del estado físico entregado aunque todavía
    # no tengan una fila patrimonial individual. Se agregan al Acta sin crear
    # componentes en Stock y evitando duplicar los ya registrados por serie.
    existing_disk_serials = {
        str(item.get("serial_number") or "").strip().upper()
        for item in groups["discos"]
        if str(item.get("serial_number") or "").strip()
    }
    existing_disk_models = {
        _normalize_component_type(item.get("brand_model") or item.get("model"))
        for item in groups["discos"]
    }
    for raw_disk in (hardware_components or {}).get("disks", []) or []:
        disk = dict(raw_disk)
        model = str(disk.get("model") or disk.get("brand_model") or "").strip()
        serial = str(disk.get("serial") or disk.get("serial_number") or "").strip()
        normalized_serial = serial.upper()
        normalized_model = _normalize_component_type(model)
        if not model:
            continue
        if normalized_serial and normalized_serial not in {"N/A", "SIN S/N", "NONE"}:
            if normalized_serial in existing_disk_serials:
                continue
        elif normalized_model in existing_disk_models:
            continue
        groups["discos"].append({
            "component_type": "Disco Rígido / SSD",
            "brand_model": model,
            "serial_number": serial if normalized_serial not in {"", "N/A", "SIN S/N", "NONE"} else None,
            "source": "telemetry",
        })
        if normalized_serial:
            existing_disk_serials.add(normalized_serial)
        existing_disk_models.add(normalized_model)

    return groups


def _has_stock_management_access():
    return (
        has_permission("funcionario")
        or has_permission("manage_stock")
        or has_permission("can_manage_stock")
    )

@bp_dashboard.route("/cementerio")
def view_cementerio():
    from utils.auth import has_permission, forbidden_response
    if not has_permission("reports"):
        return forbidden_response("reports")
    return redirect(url_for("dashboard.dashboard", estado="False"))

@bp_dashboard.route("/graficos")
def view_graphics():
    """Nueva pÃ¡gina dedicada a KPIs y GrÃ¡ficos."""
    try:
        with get_db_connection() as conn:
            kpi_total_activas = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND pc_name NOT LIKE 'PC-GENERICA%%' AND pc_name NOT LIKE 'PC GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'").fetchone()["c"]
            kpi_total_graveyard = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 0").fetchone()["c"]
            kpi_alerta_ram = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND alerta_ram_baja = 1 AND pc_name NOT LIKE 'PC-GENERICA%%' AND pc_name NOT LIKE 'PC GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'").fetchone()["c"]
            kpi_sin_impresora = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND alerta_sin_impresora = 1 AND pc_name NOT LIKE 'PC-GENERICA%%' AND pc_name NOT LIKE 'PC GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'").fetchone()["c"]
            kpi_impresora_red = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND alerta_impresora_red = 1 AND pc_name NOT LIKE 'PC-GENERICA%%' AND pc_name NOT LIKE 'PC GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'").fetchone()["c"]
            kpi_win7 = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND os_name LIKE %s AND pc_name NOT LIKE 'PC%%GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'", ("%Windows 7%",)).fetchone()["c"]
            kpi_win10 = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND os_name LIKE %s AND pc_name NOT LIKE 'PC%%GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'", ("%Windows 10%",)).fetchone()["c"]
            kpi_win11 = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND os_name LIKE %s AND pc_name NOT LIKE 'PC%%GENERICA%%' AND pc_name NOT LIKE 'INFRAESTRUCTURA%%' AND pc_name NOT LIKE 'SIGJ%%'", ("%Windows 11%",)).fetchone()["c"]
            kpi_tareas_hoy = conn.execute("SELECT COUNT(*) as c FROM tasks WHERE estado = 'Hecha' AND DATE(completed_at) = CURDATE()").fetchone()["c"]
            kpi_tareas_pendientes_count = conn.execute("SELECT COUNT(DISTINCT pc_name) as c FROM tasks WHERE estado != 'Hecha' AND pc_name IS NOT NULL AND pc_name != ''").fetchone()["c"]
            kpi_tareas_pendientes_total = conn.execute("SELECT COUNT(*) as c FROM tasks WHERE estado != 'Hecha'").fetchone()["c"]

            rows_cats = conn.execute("SELECT categoria, COUNT(*) as c FROM tasks GROUP BY categoria").fetchall()
            cat_labels = []
            cat_values = []
            for r in rows_cats:
                cat_name = r["categoria"] if r["categoria"] else "Sin CategorÃ­a"
                if cat_name == "General": continue
                cat_labels.append(cat_name)
                cat_values.append(r["c"])
            
    except Exception as e:
        print(f"Error en graficos: {e}")
        kpi_total_activas = kpi_total_graveyard = kpi_alerta_ram = kpi_sin_impresora = kpi_impresora_red = 0
        kpi_win7 = kpi_win10 = kpi_tareas_hoy = kpi_tareas_pendientes_count = kpi_tareas_pendientes_total = 0
        cat_labels = []
        cat_values = []

    return render_template(
        "graficos.html",
        kpi_total_activas=kpi_total_activas,
        kpi_total_graveyard=kpi_total_graveyard,
        kpi_alerta_ram=kpi_alerta_ram,
        kpi_sin_impresora=kpi_sin_impresora,
        kpi_impresora_red=kpi_impresora_red,
        kpi_win7=kpi_win7,
        kpi_win10=kpi_win10,
        kpi_win11=kpi_win11,
        kpi_tareas_hoy=kpi_tareas_hoy,
        kpi_tareas_pendientes_count=kpi_tareas_pendientes_count,
        kpi_tareas_pendientes_total=kpi_tareas_pendientes_total,
        cat_labels=cat_labels,
        cat_values=cat_values,
        hostname=socket.gethostname()
    )

@bp_dashboard.route("/", methods=["GET"])
def dashboard():
    """Lista todas las PCs (activas y en cementerio) + KPIs + filtros + paginado."""
    q = request.args.get("q", "").strip()
    estado = request.args.get("estado", "True").strip()
    if estado == "False":
        from utils.auth import has_permission, forbidden_response
        if not has_permission("reports"):
            return forbidden_response("reports")
    alerta = request.args.get("alerta", "").strip()
    os_param = request.args.get("os", "").strip()
    filter_tasks = request.args.get("filter_tasks", "").strip()
    sort_by = request.args.get("sort_by", "pc_name").strip()
    order = request.args.get("order", "asc").strip()
    tipo_actividad = request.args.get("tipo_actividad", "").strip()


    try: page = int(request.args.get("page", 1))
    except ValueError: page = 1
    
    try: per_page = int(request.args.get("per_page", 25))
    except ValueError: per_page = 25
    
    try:
        with get_db_connection() as conn:
            dup_row = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE alerta_nombre_duplicado = 1 AND is_active = 1").fetchone()
            duplicates_count = dup_row['c'] if dup_row else 0
            
            dup_names_rows = conn.execute("SELECT pc_name FROM pcs WHERE alerta_nombre_duplicado = 1 AND is_active = 1").fetchall()
            duplicate_pc_names = [r['pc_name'] for r in dup_names_rows]
    except Exception as e:
        print("Error checking duplicates:", e)
        duplicates_count = 0
        duplicate_pc_names = []

    template_context = load_dashboard_overview(
        q=q,
        estado=estado,
        alerta=alerta,
        os_param=os_param,
        filter_tasks=filter_tasks,
        sort_by=sort_by,
        order=order,
        page=page,
        per_page=per_page,
        tipo_actividad=tipo_actividad,
    )
    
    template_context['duplicates_count'] = duplicates_count
    template_context['duplicate_pc_names'] = duplicate_pc_names
    template_context.update(
        server_url=request.host_url,
        fuero_colors=FUERO_COLORS,
    )

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render_template("_dashboard_table_region.html", **template_context)

    return render_template(
        "index.html",
        **template_context
    )

@bp_dashboard.route("/export", methods=["GET", "POST"])
def export_inventory():
    """GET: muestra formulario. POST: genera Excel con campos seleccionados."""
    if request.method == "GET":
        return render_template("export_inventory.html")
    
    campos_seleccionados = request.form.getlist("campos")
    if not campos_seleccionados:
        # Añadimos fuero y printer_model por defecto si no hay selección
        campos_seleccionados = ["pc_name", "last_user", "fuero", "os_name", "processor", "ram_gb", "ip_address", "printer_model"]
    
    with get_db_connection() as conn:
        # Mejoramos la consulta para traer info de red e impresoras compartidas
        sql = """
            SELECT p.*,
                (SELECT GROUP_CONCAT(CONCAT(np.brand_model, ' (', np.ip_address, ')') SEPARATOR ' | ') 
                 FROM pc_network_printers pnp 
                 JOIN network_printers np ON pnp.printer_id = np.id 
                 WHERE pnp.pc_name = p.pc_name) as net_printers_info,
                (SELECT COUNT(*) FROM pcs p2 WHERE p2.is_active = 1 AND p2.printer_port LIKE CONCAT('%\\\\\\\\', p.pc_name, '%')) as clients_count
            FROM pcs p 
            WHERE p.is_active = 1 
            ORDER BY p.fuero ASC, p.pc_name ASC
        """
        rows = conn.execute(sql).fetchall()
    if not rows: return "Sin datos", 404
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Mapeo para cabeceras amigables
    headers_map = {
        "pc_name": "Nombre PC", "last_user": "Último Usuario", "fuero": "Fuero / Área",
        "os_name": "Sistema Operativo", "processor": "Procesador", "ram_gb": "RAM (GB)",
        "ip_address": "Dirección IP", "printer_model": "Impresora", "motherboard_model": "Motherboard",
        "monitors": "Monitores", "disk_models": "Discos", "last_report": "Última Sincro"
    }
    
    ws.append([headers_map.get(c, c) for c in campos_seleccionados])
    
    for row in rows:
        # Lógica para campo Impresora enriquecido
        printer_str = row["printer_model"] or "-"
        port = row["printer_port"] or ""
        
        if port.startswith("\\\\"):
            # Es una impresora compartida desde otro host
            host = port.split("\\")[2] if len(port.split("\\")) > 2 else "Red"
            printer_str = f"COMPARTIDA desde {host} ({printer_str})"
        elif row["net_printers_info"]:
            # Es una impresora de red del catálogo
            printer_str = f"RED: {row['net_printers_info']}"
        
        # Si esta PC comparte a otros
        if row["clients_count"] > 0:
            printer_str = f"Local y COMPARTIDA (Hosting a {row['clients_count']} PCs) - {printer_str}"
        elif printer_str != "-" and not printer_str.startswith(("COMPARTIDA", "RED:")):
            printer_str = f"Local: {printer_str}"

        # Creamos la fila mapeando campos
        fila = []
        for campo in campos_seleccionados:
            if campo == "printer_model":
                fila.append(printer_str)
            else:
                fila.append(row[campo])
        ws.append(fila)
    
    # --- AUTO-AJUSTE DE COLUMNAS ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2
    # -------------------------------

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Inventario_Completo_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

@bp_dashboard.route("/export_inventory_pdf", methods=["POST"])
def export_inventory_pdf():
    from services.reporting import PDFReport, format_date_es
    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT p.pc_name, p.last_user, p.fuero, p.os_name, p.printer_model, p.printer_port, p.processor, p.ram_gb, p.ip_address,
                   (SELECT GROUP_CONCAT(np.brand_model) FROM pc_network_printers pnp JOIN network_printers np ON pnp.printer_id = np.id WHERE pnp.pc_name = p.pc_name) as net_printers,
                   (SELECT COUNT(*) FROM pcs p2 WHERE p2.is_active = 1 AND p2.printer_port LIKE CONCAT('%\\\\\\\\', p.pc_name, '%')) as is_sharing_host
            FROM pcs p 
            WHERE p.is_active = 1 
              AND UPPER(p.pc_name) NOT IN ('PC GENERICA', 'INFRAESTRUCTURA', 'PC-GENERICA', 'SIGJ') 
            ORDER BY p.fuero ASC, p.pc_name ASC
        """).fetchall()
    
    pdf = PDFReport(title="Inventario Físico - Inventario GOLD", orientation='L')
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 8, f"Generado el: {format_date_es(datetime.datetime.now())}", 0, 1, 'C')
    pdf.ln(2)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Total Equipos Activos: {len(rows)}", 0, 1, 'C')
    pdf.ln(4)
    
    # Headers optimizados (A4 Landscape = 277 total útil)
    headers = ["Nombre PC", "Usuario", "Fuero/Área", "OS", "Impresora", "Procesador", "RAM", "IP Address"]
    # [PC:31, User:28, Fuero:44, OS:28, Prn:38, CPU:64, RAM:14, IP:30] = 277
    widths = [31, 28, 44, 28, 38, 64, 14, 30]
    
    pdf.set_font("Arial", "B", 8)
    pdf.set_fill_color(30, 41, 59) # Slate 800
    pdf.set_text_color(255)
    
    for i, h in enumerate(headers):
        pdf.cell(widths[i], 8, h, 1, 0, 'C', fill=True)
    pdf.ln()
    
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(0)
    
    def draw_multiline_row(pdf, cols, widths, row_height=6):
        # 1. Calcular altura necesaria (buscando el campo con más líneas)
        max_lines = 1
        split_cols = []
        for i, text in enumerate(cols):
            lines = pdf.multi_cell(widths[i], row_height, str(text), split_only=True)
            max_lines = max(max_lines, len(lines))
            split_cols.append(lines)
        
        h_row = max_lines * row_height
        
        # 2. Salto de página preventivo
        if (pdf.get_y() + h_row) > 190: 
            pdf.add_page()
            # Redibujar cabecera
            pdf.set_font("Arial", "B", 8)
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255)
            for i, h in enumerate(headers): pdf.cell(widths[i], 8, h, 1, 0, 'C', fill=True)
            pdf.ln()
            pdf.set_font("Arial", "", 8)
            pdf.set_text_color(0)

        # 3. Dibujar las celdas de la fila
        x_start, y_start = pdf.get_x(), pdf.get_y()
        for i, lines in enumerate(split_cols):
            curr_x = x_start + sum(widths[:i])
            pdf.set_xy(curr_x, y_start)
            # Dibujar borde de la celda
            pdf.rect(curr_x, y_start, widths[i], h_row)
            cell_text = "\n".join(lines)
            pdf.multi_cell(widths[i], row_height, cell_text, 0, 'L')
        
        pdf.set_xy(x_start, y_start + h_row)

    for row in rows:
        # Limpieza de datos
        raw_user = row["last_user"] or "N/A"
        user = raw_user.split("\\")[-1] if "\\" in raw_user else raw_user
        
        os_str = (row["os_name"] or "N/A").replace("Microsoft ", "")
        
        # --- Lógica de Impresora Detallada ---
        printer = row["printer_model"] or "-"
        port = row["printer_port"] or ""
        
        if port.startswith("\\\\"):
            # Caso: Impresora en red compartida por otra PC
            host_srv = port.split("\\")[2] if len(port.split("\\")) > 2 else "Server"
            printer = f"Compartida (desde {host_srv})"
        elif row["net_printers"]:
            # Caso: Impresora de red directa (Catálogo)
            printer = f"Red ({row['net_printers']})"
        elif printer.upper() in ("N/A", "SIN IMPRESORA", "NONE", "-"):
            printer = "-"
        else:
            # Es local, ver si la comparte
            if row["is_sharing_host"] > 0:
                printer = f"Local y Compartida (Hosting a {row['is_sharing_host']} PCs) - {printer}"
            else:
                printer = f"Local ({printer})"
        # -------------------------------------
        
        data_to_draw = [
            str(row["pc_name"]),
            str(user),
            str(row["fuero"] or "N/A"),
            str(os_str),
            str(printer),
            str(row["processor"] or "N/A"),
            f'{row["ram_gb"]}G',
            str(row["ip_address"])
        ]
        
        draw_multiline_row(pdf, data_to_draw, widths)

    output = BytesIO()
    pdf_bytes = pdf.output()
    output.write(pdf_bytes)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"Inventario_Fisico_{dt.now().strftime('%Y%m%d')}.pdf",
    )

@bp_dashboard.route("/download_db")
def download_db():
    return "Backup de BD no disponible en modo MySQL. Use mysqldump desde el servidor.", 503

@bp_dashboard.route("/decommission/<string:pc_name>", methods=["POST"])
def decommission_pc(pc_name):
    """Pasar una PC al cementerio."""
    if decommission_pc_service(pc_name, request.remote_addr):
        return redirect(url_for("dashboard.dashboard"))
    return "Error al dar de baja", 500

@bp_dashboard.route("/reactivate/<pc_name>", methods=["POST"])
def reactivate_pc(pc_name):
    """Reactivar una PC."""
    if reactivate_pc_service(pc_name, request.remote_addr):
        return redirect(url_for("dashboard.dashboard"))
    return "Error al reactivar", 500

@bp_dashboard.route("/refresh_fueros", methods=["POST"])
def refresh_fueros():
    """Recalcula el fuero para todas las PCs basÃ¡ndose en el nombre."""
    try:
        with get_db_connection() as conn:
            result = recalculate_all_pc_fueros(conn)
        print(f"Fueros actualizados para {result['updated']} PCs.")
    except Exception as exc:
        print(f"Error refreshing fueros: {exc}")

    return redirect(url_for("dashboard.dashboard"))

@bp_dashboard.route("/pc/<pc_name>/update_fuero", methods=["POST"])
def update_pc_fuero(pc_name):
    fuero = request.form.get("fuero", "").strip()
    try:
        with get_db_connection() as conn:
            old_pc = conn.execute("SELECT fuero FROM pcs WHERE pc_name = %s", (pc_name,)).fetchone()
            if not old_pc:
                abort(404)
            old_fuero = old_pc["fuero"] or ""
            conn.execute("UPDATE pcs SET fuero = %s WHERE pc_name = %s", (fuero or None, pc_name))
            if old_fuero != fuero:
                log_audit_event(
                    conn,
                    pc_name=pc_name,
                    field="fuero",
                    old_value=old_fuero,
                    new_value=fuero,
                    action_type="EDICION_FUERO",
                    request_ip=request.remote_addr,
                )
            conn.commit()
        flash(f"Fuero actualizado para {pc_name}.", "success")
    except Exception as exc:
        flash(f"No se pudo actualizar el fuero de {pc_name}: {exc}", "error")
    return redirect(request.referrer or url_for("dashboard.dashboard"))

@bp_dashboard.route("/delete_permanent/<string:pc_name>", methods=["POST"])
def delete_permanent_pc(pc_name):
    """Borrado definitivo de una PC y sus tareas asociadas."""
    if delete_permanent_pc_service(pc_name, request.remote_addr):
        return redirect(url_for("dashboard.dashboard"))
    return "Error al borrar permanentemente", 500



def generate_next_bo_code(conn):
    """Genera el siguiente código BO-YYYY-XXXX garantizando que sea único y evitando duplicados."""
    year = datetime.datetime.now().strftime("%Y")
    prefix = f"BO-{year}-"
    rows = conn.execute(
        "SELECT code FROM build_orders WHERE code LIKE %s",
        (f"{prefix}%",)
    ).fetchall()

    max_seq = 0
    for r in rows:
        code_str = r.get("code") or ""
        try:
            num = int(code_str.replace(prefix, ""))
            if num > max_seq:
                max_seq = num
        except ValueError:
            pass

    seq = max_seq + 1
    while True:
        candidate = f"BO-{year}-{seq:04d}"
        exists = conn.execute("SELECT id FROM build_orders WHERE code = %s", (candidate,)).fetchone()
        if not exists:
            return candidate
        seq += 1


@bp_dashboard.route("/pc/<pc_name>/create_bo_from_telemetry", methods=["POST"])
@login_required
def create_bo_from_telemetry(pc_name):
    """Crea la primera orden o agrega cambios a la orden patrimonial existente."""
    import logging
    logger = logging.getLogger(__name__)
    if not _has_stock_management_access():
        flash("Acceso denegado: se requiere permiso de Gestión Stock.", "error")
        return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))
    try:
        from utils.auth import current_technician_identity
        tech = current_technician_identity() or "Sistema"
        target_user = request.form.get("target_user", "").strip() or None
        target_fuero = request.form.get("target_fuero", "").strip() or None
        invoice_number = request.form.get("invoice_number", "").strip() or None
        oc_number = request.form.get("oc_number", "").strip() or None
        notes = request.form.get("notes", "").strip() or "Generada desde Telemetría (.ps1)"
        
        comp_selected = request.form.getlist("comp_selected")
        comp_types = request.form.getlist("comp_type")
        comp_models = request.form.getlist("comp_model")
        comp_serials = request.form.getlist("comp_serial")
        comp_invoices = request.form.getlist("comp_invoice")
        comp_ocs = request.form.getlist("comp_oc")
        
        selected_indices = set()
        for idx_str in comp_selected:
            try:
                selected_indices.add(int(idx_str))
            except ValueError:
                pass

        # Fallback si no vinieron índices seleccionados explicitamente
        if not selected_indices and comp_types:
            selected_indices = set(range(len(comp_types)))

        with get_db_connection() as conn:
            clean_name = pc_name.strip().lower()
            pc = conn.execute(
                """
                SELECT pc_name, last_user, fuero, validation_status
                FROM pcs
                WHERE LOWER(TRIM(pc_name)) = %s
                FOR UPDATE
                """,
                (clean_name,),
            ).fetchone()
            pc_dict = dict(pc) if pc else {
                "pc_name": pc_name,
                "last_user": None,
                "fuero": None,
                "validation_status": "sin_gemelo",
            }

            final_user = target_user or pc_dict.get("last_user")
            final_fuero = target_fuero or pc_dict.get("fuero")

            # Fallback de O.C. y Remito General desde los componentes individuales si la cabecera vino vacía
            if not oc_number:
                first_oc = next((oc.strip() for idx, oc in enumerate(comp_ocs) if idx in selected_indices and oc and oc.strip()), None)
                if first_oc:
                    oc_number = first_oc

            if not invoice_number:
                first_inv = next((inv.strip() for idx, inv in enumerate(comp_invoices) if idx in selected_indices and inv and inv.strip()), None)
                if first_inv:
                    invoice_number = first_inv

            # Validar si ya existe una Orden de Armado previa para esta PC (Actualizar o Crear)
            existing_bo = conn.execute(
                """
                SELECT DISTINCT bo.id, bo.code
                FROM build_orders bo
                LEFT JOIN build_order_items boi ON boi.build_order_id = bo.id
                WHERE LOWER(TRIM(bo.target_pc_name)) = %s
                   OR LOWER(TRIM(boi.pc_name)) = %s
                ORDER BY bo.created_at DESC
                LIMIT 1
                """,
                (clean_name, clean_name),
            ).fetchone()

            # Un gemelo ya confirmado puede provenir de una asignación directa
            # antigua, sin Orden de Armado. No fabricamos una orden nueva al
            # volver a recibir el mismo reporte: esa ejecución queda en historial.
            current_validation = (pc_dict.get("validation_status") or "sin_gemelo").strip().lower()
            existing_component = None
            if not existing_bo and current_validation == "sin_gemelo":
                existing_component = conn.execute(
                    """
                    SELECT id
                    FROM components
                    WHERE LOWER(TRIM(assigned_pc)) = %s
                      AND (status IS NULL OR status NOT IN ('Retirado', 'Scrap', 'Stock'))
                      AND (lifecycle_status IS NULL OR lifecycle_status NOT IN ('retirado', 'scrap', 'stock'))
                    LIMIT 1
                    """,
                    (clean_name,),
                ).fetchone()

            if not existing_bo and (current_validation != "sin_gemelo" or existing_component):
                flash(
                    "El equipo ya posee un gemelo patrimonial. El nuevo reporte queda en su historial; no se creó otra Orden de Armado.",
                    "info",
                )
                return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))

            updating_existing = bool(existing_bo)
            if existing_bo:
                bo_id = existing_bo["id"]
                code = existing_bo["code"]
                conn.execute(
                    """
                    UPDATE build_orders 
                    SET oc_number = COALESCE(%s, oc_number),
                        invoice_number = COALESCE(%s, invoice_number),
                        target_fuero = COALESCE(%s, target_fuero),
                        target_user = COALESCE(%s, target_user),
                        target_pc_name = %s,
                        notes = %s,
                        status = 'completed'
                    WHERE id = %s
                    """,
                    (oc_number, invoice_number, final_fuero, final_user, pc_name, notes, bo_id)
                )
            else:
                code = generate_next_bo_code(conn)
                conn.execute(
                    """
                    INSERT INTO build_orders (code, oc_number, invoice_number, target_fuero, target_user, target_pc_name, notes, created_by, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'completed')
                    """,
                    (code, oc_number, invoice_number, final_fuero, final_user, pc_name, notes, tech)
                )
                bo_id = conn.cursor.lastrowid

            for idx in range(len(comp_types)):
                if idx not in selected_indices:
                    continue

                c_type = comp_types[idx].strip()
                c_model = comp_models[idx].strip()
                c_serial = comp_serials[idx].strip()
                item_inv = comp_invoices[idx].strip() if idx < len(comp_invoices) and comp_invoices[idx].strip() else invoice_number
                item_oc = comp_ocs[idx].strip() if idx < len(comp_ocs) and comp_ocs[idx].strip() else oc_number

                if not c_type or not c_model:
                    continue

                clean_sn = c_serial if c_serial and c_serial.upper() not in ("N/A", "SIN S/N", "NONE", "") else None
                
                comp_id = None
                if clean_sn:
                    existing = conn.execute(
                        "SELECT id, serial_number FROM components WHERE UPPER(serial_number) = %s LIMIT 1",
                        (clean_sn.upper(),),
                    ).fetchone()
                    if existing:
                        comp_id = existing["id"]
                        clean_sn = existing.get("serial_number") or clean_sn
                        deployed_status, deployed_lifecycle = deployed_component_state()
                        conn.execute(
                            """
                            UPDATE components 
                            SET build_order_id = %s, assigned_pc = %s, status = %s, lifecycle_status = %s, assigned_user = %s, assigned_fuero = %s,
                                invoice_number = COALESCE(%s, invoice_number), oc_number = COALESCE(%s, oc_number)
                            WHERE id = %s
                            """,
                            (bo_id, pc_name, deployed_status, deployed_lifecycle, final_user, final_fuero, item_inv, item_oc, comp_id)
                        )

                # Componentes WMI sin serie (CPU/RAM) se reconocen por tipo y
                # modelo dentro del mismo gemelo. Así un reporte repetido no
                # genera otro Auto-ID.
                if not comp_id and not clean_sn:
                    existing = conn.execute(
                        """
                        SELECT id, serial_number
                        FROM components
                        WHERE (LOWER(TRIM(assigned_pc)) = %s OR build_order_id = %s)
                          AND LOWER(TRIM(component_type)) = LOWER(TRIM(%s))
                          AND LOWER(TRIM(brand_model)) = LOWER(TRIM(%s))
                          AND status NOT IN ('Retirado', 'Scrap')
                          AND (lifecycle_status IS NULL OR lifecycle_status NOT IN ('retirado', 'scrap'))
                        LIMIT 1
                        """,
                        (clean_name, bo_id, c_type, c_model),
                    ).fetchone()
                    if existing:
                        comp_id = existing["id"]
                        clean_sn = existing.get("serial_number")

                if not comp_id:
                    if clean_sn:
                        sn_to_insert = clean_sn
                    else:
                        from blueprints.bp_stock import generate_internal_serial
                        sn_to_insert = generate_internal_serial(c_type)
                    deployed_status, deployed_lifecycle = deployed_component_state()
                    conn.execute(
                        """
                        INSERT INTO components (serial_number, component_type, brand_model, status, lifecycle_status, assigned_pc, build_order_id, assigned_user, assigned_fuero, invoice_number, oc_number)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (sn_to_insert, c_type, c_model, deployed_status, deployed_lifecycle, pc_name, bo_id, final_user, final_fuero, item_inv, item_oc)
                    )
                    clean_sn = sn_to_insert

                existing_item = conn.execute(
                    """
                    SELECT id FROM build_order_items
                    WHERE build_order_id = %s AND UPPER(TRIM(serial_number)) = UPPER(TRIM(%s))
                    LIMIT 1
                    """,
                    (bo_id, clean_sn),
                ).fetchone()
                if not existing_item:
                    conn.execute(
                        """
                        INSERT INTO build_order_items (build_order_id, serial_number, asset_type, brand_model, pc_name, scanned_by)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (bo_id, clean_sn, c_type, c_model, pc_name, tech)
                    )

            conn.execute("UPDATE pcs SET validation_status = 'validado' WHERE pc_name = %s", (pc_name,))

            log_audit_event(
                conn,
                pc_name=pc_name,
                field="build_order",
                old_value=code if updating_existing else None,
                new_value=code,
                action_type="ACTUALIZAR_BO_TELEMETRIA" if updating_existing else "CREAR_BO_TELEMETRIA",
                request_ip=request.remote_addr,
            )
            conn.commit()

        if updating_existing:
            flash(f"Se actualizó la Orden de Armado {code} sin duplicar sus componentes ni borrar su historial.", "success")
        else:
            flash(f"Se creó con éxito la Orden de Armado {code} con los componentes seleccionados.", "success")
    except Exception as exc:
        logger.error("Error creando Orden de Armado desde telemetría para %s: %s", pc_name, exc)
        flash(f"Error al generar la Orden de Armado: {exc}", "error")

    return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))


@bp_dashboard.route("/pc/<pc_name>")
def pc_detail(pc_name):
    """Detalle de una PC."""
    ctx = get_pc_detail_context(pc_name)
    if not ctx:
        from flask import abort
        abort(404)
    return render_template("pc_detail.html", **ctx, fuero_colors=FUERO_COLORS)

@bp_dashboard.route("/pc/<pc_name>/acta_gemelo_validado")
def acta_gemelo_validado(pc_name):
    """Genera la plantilla imprimible del Acta de Entrega y Conformidad (Gemelo Validado OK)."""
    from datetime import datetime

    ctx = get_pc_detail_context(pc_name)
    if not ctx or not ctx.get("pc"):
        from flask import abort
        abort(404)

    pc = ctx["pc"]
    val_status = pc.get("validation_status")
    validation_comp = ctx.get("validation_comparison") or []
    has_discrepancy = any(not item.get("match") for item in validation_comp)

    # REGLA DE NEGOCIO RESTRICTIVA: Solo disponible si Gemelo Digital está Validado OK y sin discrepancias
    if val_status != "validado" or has_discrepancy:
        flash("El Acta de Entrega sólo está disponible para equipos con Gemelo Digital Validado OK.", "warning")
        return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))

    components = list(
        ctx.get("display_components")
        or ctx.get("all_unified_components")
        or ctx.get("pc_components")
        or ctx.get("components")
        or []
    )
    # El lector USB se ignora para validar el gemelo, pero el Acta debe enumerar
    # todo el almacenamiento físico reportado/registrado, incluido el removible.
    known_serials = {
        str(item.get("serial_number") or "").strip().upper()
        for item in components
        if item
    }
    for item in ctx.get("all_unified_components") or []:
        serial = str(item.get("serial_number") or "").strip().upper()
        if is_ignored_storage_component(item) and serial not in known_serials:
            components.append(dict(item))
            if serial:
                known_serials.add(serial)
    monitors_detail = ctx.get("display_monitors_detail") or ctx.get("monitors_detail") or []
    acta_component_groups = build_acta_component_groups(
        components,
        monitors_detail,
        ctx.get("hardware_components"),
    )

    # Identificar técnico conectado
    auth_user = session.get("auth_user") or {}
    tecnico_user = auth_user.get("full_name") or auth_user.get("username") or "Departamento de Informática"

    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M hs")

    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    server_host = request.host
    if any(h in server_host for h in ["localhost", "127.0.0.1", "0.0.0.0"]):
        server_host = os.environ.get("SERVER_PUBLIC_HOST", "10.15.2.251:5000")
    qr_url = f"{scheme}://{server_host}/public/asset/{pc_name}"

    return render_template(
        "acta_entrega_gemelo.html",
        pc=pc,
        linked_bo=ctx.get("linked_bo"),
        components=components,
        monitors_detail=monitors_detail,
        acta_component_groups=acta_component_groups,
        tecnico_user=tecnico_user,
        generated_at=generated_at,
        qr_url=qr_url
    )


@bp_dashboard.route("/api/pc/<pc_name>/resolve_autogenerated_duplicates", methods=["POST"])
@login_required
def resolve_autogenerated_duplicates(pc_name):
    if not _has_stock_management_access():
        return jsonify({"status": "error", "message": "Acceso denegado: se requiere permiso de Gestión Stock."}), 403

    ctx = get_pc_detail_context(pc_name)
    candidates = ctx.get("autogenerated_shadow_candidates") or []
    serials = [c.get("serial_number") for c in candidates if c.get("serial_number")]

    if not serials:
        return jsonify({"status": "success", "deleted_count": 0, "deleted_serials": [], "message": "No se detectaron Auto-ID duplicados para resolver."})

    placeholders = ",".join(["%s"] * len(serials))
    actor = current_technician_identity() or current_username() or "Sistema"
    deleted_serials = []

    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT serial_number, component_type, brand_model, assigned_pc
            FROM components
            WHERE serial_number IN ({placeholders})
            """,
            tuple(serials),
        ).fetchall()

        for row in rows:
            component = dict(row)
            serial = component.get("serial_number")
            old_pc = component.get("assigned_pc") or pc_name
            conn.execute("DELETE FROM components WHERE serial_number = %s", (serial,))
            conn.execute(
                """
                INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    old_pc,
                    "AUTO_ID_DUPLICATE_RESOLVED",
                    f"{component.get('component_type')} {component.get('brand_model')} (S/N: {serial})",
                    "DELETED",
                    actor,
                    "AUTO_ID_DUPLICATE_RESOLVED",
                    request.remote_addr,
                ),
            )
            deleted_serials.append(serial)

    return jsonify({
        "status": "success",
        "deleted_count": len(deleted_serials),
        "deleted_serials": deleted_serials,
        "message": f"Se resolvieron {len(deleted_serials)} Auto-ID duplicado(s) en {pc_name}.",
    })


@bp_dashboard.route("/public/asset/<pc_name>")
def public_asset_info(pc_name):
    """Vista pública/móvil optimizada para lectura de QR desde smartphone (Soporta PC o S/N Componente)."""
    try:
        from services.pc_details_service import get_pc_detail_context
        
        # 1. Probar si es una PC registrada en pcs
        ctx = get_pc_detail_context(pc_name)
        if ctx and ctx.get("pc"):
            all_comps = ctx.get("display_components") or ctx.get("all_unified_components") or ctx.get("pc_components") or ctx.get("components") or []
            mon_detail = list(ctx.get("display_monitors_detail") or ctx.get("monitors_detail") or [])
            if not mon_detail:
                mon_detail = [
                    c for c in all_comps 
                    if "MONITOR" in (c.get("component_type") or "").strip().upper() 
                    or "MONITOR" in (c.get("brand_model") or "").strip().upper()
                ]
            ctx["components"] = all_comps
            ctx["monitors_detail"] = mon_detail
            ctx["is_standalone_component"] = False
            ctx["is_authenticated"] = is_authenticated()
            return render_template("public_asset_info.html", **ctx)
        
        # 2. Si no es una PC directamente, buscar en componentes por Número de Serie
        with get_db_connection() as conn:
            comp_row = conn.execute(
                """
                SELECT c.*, bo.code as bo_code
                FROM components c
                LEFT JOIN build_orders bo ON bo.id = c.build_order_id
                WHERE LOWER(c.serial_number) = %s OR LOWER(c.serial_number) LIKE %s
                LIMIT 1
                """,
                (pc_name.lower(), f"%{pc_name.lower()}%")
            ).fetchone()
            
            if comp_row:
                c_dict = dict(comp_row)
                assigned_pc = c_dict.get("assigned_pc")
                if not assigned_pc and c_dict.get("build_order_id"):
                    bo = conn.execute("SELECT target_pc_name FROM build_orders WHERE id = %s", (c_dict["build_order_id"],)).fetchone()
                    if bo and bo.get("target_pc_name"):
                        assigned_pc = bo["target_pc_name"]

                if assigned_pc:
                    pc_ctx = get_pc_detail_context(assigned_pc)
                    if pc_ctx and pc_ctx.get("pc"):
                        all_comps = pc_ctx.get("display_components") or pc_ctx.get("all_unified_components") or pc_ctx.get("pc_components") or pc_ctx.get("components") or []
                        mon_detail = list(pc_ctx.get("display_monitors_detail") or pc_ctx.get("monitors_detail") or [])
                        if not mon_detail:
                            mon_detail = [
                                c for c in all_comps 
                                if "MONITOR" in (c.get("component_type") or "").strip().upper() 
                                or "MONITOR" in (c.get("brand_model") or "").strip().upper()
                            ]
                        pc_ctx["components"] = all_comps
                        pc_ctx["monitors_detail"] = mon_detail
                        pc_ctx["scanned_comp"] = c_dict
                        pc_ctx["is_standalone_component"] = False
                        pc_ctx["is_authenticated"] = is_authenticated()
                        return render_template("public_asset_info.html", **pc_ctx)
                
                fake_pc = {
                    "pc_name": c_dict.get("serial_number"),
                    "fuero": c_dict.get("assigned_fuero") or "Depósito / Stock",
                    "last_user": c_dict.get("assigned_user") or "En Stock",
                    "validation_status": "validado" if c_dict.get("status") in ["Stock", "Installed"] else "pendiente"
                }
                return render_template(
                    "public_asset_info.html",
                    pc_name=c_dict.get("serial_number"),
                    pc=fake_pc,
                    components=[c_dict],
                    monitors_detail=[c_dict] if (c_dict.get("component_type") or "").upper() == "MONITOR" else [],
                    is_standalone_component=True,
                    is_authenticated=is_authenticated()
                )

    except Exception as e:
        import logging
        logging.error("Error en public_asset_info para %s: %s", pc_name, e)

    return render_template("public_asset_info.html", pc_name=pc_name, pc={}, components=[], monitors_detail=[], is_authenticated=False)

@bp_dashboard.route("/pc/<pc_name>/qr_label")
def pc_qr_label_view(pc_name):
    """Renderiza la plantilla de doble etiqueta QR (Gabinete CPU + Monitor)."""
    ctx = get_pc_detail_context(pc_name)
    if not ctx:
        from flask import abort
        abort(404)
    
    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    server_host = request.host
    if any(h in server_host for h in ["localhost", "127.0.0.1", "0.0.0.0"]):
        server_host = os.environ.get("SERVER_PUBLIC_HOST", "10.15.2.251:5000")
    qr_url = f"{scheme}://{server_host}/public/asset/{pc_name}"
    
    components = ctx.get("display_components") or ctx.get("all_unified_components") or ctx.get("pc_components") or ctx.get("components") or []
    monitors_detail = ctx.get("display_monitors_detail") or ctx.get("monitors_detail") or []
    
    monitor_comp = next((c for c in components if (c.get("component_type") or "").upper() == "MONITOR"), None)
    if not monitor_comp and monitors_detail:
        first_mon = monitors_detail[0]
        monitor_comp = {
            "component_type": "MONITOR",
            "brand_model": first_mon.get("brand_model"),
            "serial_number": first_mon.get("serial_number"),
            "invoice_number": first_mon.get("invoice_number"),
            "oc_number": first_mon.get("oc_number"),
            "supplier": first_mon.get("supplier")
        }
        
    cpu_comp = next((c for c in components if (c.get("component_type") or "").upper() in ["GABINETE", "CPU", "MOTHERBOARD"]), None)
    keyboard_comp = next((c for c in components if (c.get("component_type") or "").upper() == "TECLADO"), None)
    mouse_comp = next((c for c in components if (c.get("component_type") or "").upper() == "MOUSE"), None)
    
    return render_template(
        "qr_double_label.html",
        pc_name=pc_name,
        pc_info=ctx.get("pc", {}),
        components=components,
        monitors_detail=monitors_detail,
        monitor_comp=monitor_comp,
        cpu_comp=cpu_comp,
        keyboard_comp=keyboard_comp,
        mouse_comp=mouse_comp,
        qr_url=qr_url,
        server_host=server_host
    )


@bp_dashboard.route("/global_activity")
@bp_dashboard.route("/actividad_global")
def global_activity():
    """Muestra la vista de Auditoría y Historial Global de Actividad."""
    from utils.auth import has_permission, forbidden_response
    if not has_permission("dashboard") and not has_permission("reports"):
        return forbidden_response("dashboard")
    try:
        with get_db_connection() as conn:
            logs = conn.execute(
                "SELECT * FROM audit_logs ORDER BY changed_at DESC LIMIT 500"
            ).fetchall()
        return render_template("activity_logs.html", logs=logs)
    except Exception as e:
        flash(f"Error al cargar el historial de auditoría: {e}", "danger")
        return redirect(url_for("dashboard.dashboard"))


@bp_dashboard.route("/pc/<pc_name>/update_infrastructure", methods=["POST"])
def update_pc_infrastructure(pc_name):
    """Actualiza datos de red y ubicación de una PC desde su vista de detalle."""
    infra_data = {
        'building': request.form.get('building', '').strip(),
        'floor': request.form.get('floor', '').strip(),
        'switch_name': request.form.get('switch_name', '').strip(),
        'switch_port': request.form.get('switch_port', '').strip(),
        'pachera_name': request.form.get('pachera_name', '').strip(),
        'pachera_port': request.form.get('pachera_port', '').strip(),
    }
    if update_pc_infrastructure_service(pc_name, infra_data, request.remote_addr):
        flash(f"Infraestructura de {pc_name} actualizada.", "success")
    else:
        flash(f"Error al actualizar infraestructura de {pc_name}.", "error")
    return redirect(request.referrer or url_for("dashboard.pc_detail", pc_name=pc_name))


@bp_dashboard.route("/pc/<pc_name>/update_serials", methods=["POST"])
def update_pc_serials(pc_name):
    """Actualiza los números de serie de CPU/Gabinete, monitor, impresora y motherboard de una PC."""
    cpu_sn = request.form.get('cpu_sn', '').strip()
    monitor_sn = request.form.get('monitor_sn', '').strip()
    printer_sn = request.form.get('printer_sn', '').strip()
    motherboard_sn = request.form.get('motherboard_sn', '').strip()
    try:
        with get_db_connection() as conn:
            conn.execute(
                "UPDATE pcs SET serial_number = %s, serial_monitor = %s, serial_impresora = %s WHERE pc_name = %s",
                (cpu_sn or None, monitor_sn or None, printer_sn or None, pc_name)
            )
            if motherboard_sn:
                conn.execute(
                    "UPDATE components SET serial_number = %s WHERE LOWER(TRIM(assigned_pc)) = LOWER(TRIM(%s)) AND LOWER(component_type) IN ('motherboard', 'placa madre', 'mother')",
                    (motherboard_sn, pc_name)
                )
            conn.commit()
        flash(f"Números de serie de {pc_name} actualizados correctamente.", "success")
    except Exception as e:
        flash(f"Error al actualizar números de serie: {e}", "error")
    return redirect(request.referrer or url_for("dashboard.pc_detail", pc_name=pc_name))
