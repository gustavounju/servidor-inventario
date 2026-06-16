from flask import Blueprint, request, jsonify, redirect, url_for, render_template, send_file
import datetime
from datetime import datetime as dt
import math
import uuid
import pymysql
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.db_core import get_db_connection
from services.ai_assistant import predict_category
from services.reporting import PDFReport, format_datetime_es, format_date_es
from services.push_notifications import notify_all_technicians
from utils.auth import superuser_required, current_username, list_technician_users

bp_tasks = Blueprint('tasks', __name__)
LOCAL_UTC_OFFSET_HOURS = -3


def _normalize_directory_username(value):
    cleaned = (value or "").strip().lower()
    if "\\" in cleaned:
        cleaned = cleaned.split("\\")[-1]
    if "@" in cleaned:
        cleaned = cleaned.split("@")[0]
    return cleaned


def _is_generic_pc_name(pc_name):
    normalized = "".join(ch for ch in (pc_name or "").upper() if ch.isalpha())
    return "PCGENERICA" in normalized


def _build_task_user_match_index(conn):
    ad_rows = conn.execute(
        "SELECT username, real_name FROM ad_users"
    ).fetchall()
    real_name_to_username = {}
    known_usernames = set()
    for row in ad_rows:
        username = _normalize_directory_username(row.get("username"))
        real_name = (row.get("real_name") or "").strip().lower()
        if username:
            known_usernames.add(username)
        if username and real_name and real_name not in real_name_to_username:
            real_name_to_username[real_name] = username

    pc_rows = conn.execute(
        """
        SELECT pc_name, last_user, fuero
        FROM pcs
        WHERE is_active = 1
        ORDER BY pc_name
        """
    ).fetchall()
    pcs_by_username = {}
    for row in pc_rows:
        pc_name = row.get("pc_name") or ""
        if not pc_name or _is_generic_pc_name(pc_name) or pc_name.upper().startswith("INFRAESTRUCTURA"):
            continue
        username = _normalize_directory_username(row.get("last_user"))
        if not username:
            continue
        pcs_by_username.setdefault(username, []).append(
            {
                "pc_name": pc_name,
                "last_user": row.get("last_user") or "",
                "fuero": row.get("fuero") or "",
            }
        )

    return real_name_to_username, known_usernames, pcs_by_username


def _attach_task_user_match(task, match_index):
    real_name_to_username, known_usernames, pcs_by_username = match_index
    solicitante = (task.get("solicitante") or "").strip()
    solicitante_key = solicitante.lower()
    matched_username = ""
    direct_username = _normalize_directory_username(solicitante)
    if direct_username and (direct_username in known_usernames or direct_username in pcs_by_username):
        matched_username = direct_username
    elif solicitante_key in real_name_to_username:
        matched_username = real_name_to_username[solicitante_key]
    elif len(solicitante_key) >= 4:
        # Intento de coincidencia parcial (fuzzy match)
        for rn_key, uname in real_name_to_username.items():
            if solicitante_key in rn_key or rn_key in solicitante_key:
                matched_username = uname
                break

    matched_pcs = []
    current_pc = task.get("pc_name") or ""
    current_pc_upper = current_pc.upper()
    if matched_username:
        for candidate in pcs_by_username.get(matched_username, []):
            matched_pcs.append(
                {
                    **candidate,
                    "is_current_pc": candidate["pc_name"].upper() == current_pc_upper,
                }
            )

    task["matched_username"] = matched_username
    task["matched_pcs"] = matched_pcs
    task["matched_pc_count"] = len(matched_pcs)
    task["has_user_match"] = bool(matched_pcs)
    task["suggested_pc_name"] = matched_pcs[0]["pc_name"] if len(matched_pcs) == 1 else ""
    return task


def _attach_task_user_matches(task_rows, conn):
    match_index = _build_task_user_match_index(conn)
    enriched = []
    for row in task_rows:
        enriched.append(_attach_task_user_match(dict(row), match_index))
    return enriched


def _coerce_task_datetime_for_display(dt_value, now=None):
    if not dt_value:
        return dt_value
    now = now or dt.now()
    # Corrige registros heredados guardados en UTC para el visor local (-03:00).
    if dt_value > now + datetime.timedelta(hours=2):
        return dt_value + datetime.timedelta(hours=LOCAL_UTC_OFFSET_HOURS)
    return dt_value

def _is_assignable_technician(name):
    target = (name or "").strip().lower()
    if not target:
        return False
    return any((tech.get("name") or "").strip().lower() == target for tech in list_technician_users())

def _format_duration(start, end):
    if not start:
        return ""
    end = end or dt.now()
    total_minutes = max(0, int((end - start).total_seconds() // 60))
    if total_minutes < 1:
        return "<1m"
    days, rem = divmod(total_minutes, 1440)
    hours, minutes = divmod(rem, 60)
    if days:
        return f"{days}d {hours}h"
    if hours:
        return f"{hours}h {minutes}m"
    return f"{minutes}m"

def _decorate_visor_task(row):
    task = dict(row)
    created_at = _coerce_task_datetime_for_display(task.get("created_at"))
    completed_at = _coerce_task_datetime_for_display(task.get("completed_at"))
    task["created_at"] = created_at
    task["completed_at"] = completed_at
    task["created_at_fmt"] = created_at.strftime("%d/%m %H:%M") if created_at else ""
    task["created_at_time"] = created_at.strftime("%H:%M") if created_at else ""
    task["completed_at_fmt"] = completed_at.strftime("%d/%m %H:%M") if completed_at else ""
    task["resolution_time"] = _format_duration(created_at, completed_at if task.get("estado") == "Hecha" else None)
    task["resolution_label"] = "Resolución" if task.get("estado") == "Hecha" else "Abierta"
    if "has_user_match" not in task:
        task["has_user_match"] = False
    if "matched_pcs" not in task:
        task["matched_pcs"] = []
    if "matched_pc_count" not in task:
        task["matched_pc_count"] = 0
    if "matched_username" not in task:
        task["matched_username"] = ""
    return task

@bp_tasks.app_context_processor
def inject_tasks_kpis():
    from utils.auth import allowed_module_links, auth_mode_label, available_roles, current_user, has_permission, is_authenticated, role_label, generate_csrf_token
    from utils.constants import APP_VERSION
    from utils.runtime_urls import get_public_app_base_url, get_public_script_fallback_url
    
    kpis = {}
    if is_authenticated():
        try:
            with get_db_connection() as conn:
                kpis['kpi_total_activas'] = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND UPPER(pc_name) NOT IN ('PC GENERICA', 'INFRAESTRUCTURA', 'PC-GENERICA')").fetchone()["c"]
                kpis['kpi_total_graveyard'] = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 0").fetchone()["c"]
                kpis['kpi_win7'] = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND os_name LIKE %s AND UPPER(pc_name) NOT IN ('PC GENERICA', 'INFRAESTRUCTURA', 'PC-GENERICA')", ("%Windows 7%",)).fetchone()["c"]
                kpis['kpi_alerta_ram'] = conn.execute("SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 AND alerta_ram_baja = 1 AND UPPER(pc_name) NOT IN ('PC GENERICA', 'INFRAESTRUCTURA', 'PC-GENERICA')").fetchone()["c"]
                net_pr = conn.execute("SELECT COUNT(*) as c FROM network_printers").fetchone()["c"]
                loc_pr = conn.execute("""
                    SELECT COUNT(*) as c FROM pcs WHERE is_active = 1 
                    AND (printer_model IS NOT NULL AND printer_model != '' AND printer_model != 'N/A' AND UPPER(printer_model) NOT LIKE '%%SIN IMPRESORA%%')
                    AND (printer_port IS NULL OR printer_port NOT LIKE '\\\\\\\\%%') AND alerta_impresora_red = 0
                    AND pc_name NOT IN (SELECT pc_name FROM pc_network_printers)
                """).fetchone()["c"]
                kpis['kpi_total_impresoras'] = net_pr + loc_pr
                kpis['kpi_tareas_hoy'] = conn.execute("SELECT COUNT(*) as c FROM tasks WHERE estado = 'Hecha' AND DATE(completed_at) = CURDATE()").fetchone()["c"]
                # Mejoramos la consistencia del conteo de pendientes
                kpis['kpi_total_pendientes'] = conn.execute("SELECT COUNT(*) as c FROM tasks WHERE estado != 'Hecha'").fetchone()["c"]
                # Usuarios pendientes de aprobación (AD)
                kpis['kpi_usuarios_pendientes'] = conn.execute("SELECT COUNT(*) as c FROM app_users WHERE is_active = 0").fetchone()["c"]
        except Exception as e:
            print(f"Error in context processor KPIs (Tasks): {e}")

    return {
        'app_version': APP_VERSION,
        'csrf_token': generate_csrf_token,
        'is_authenticated': is_authenticated(),
        'current_user': current_user(),
        'auth_mode_label': auth_mode_label(),
        'has_access': has_permission,
        'module_access_links': allowed_module_links(),
        'current_role_label': role_label(),
        'available_roles': available_roles(),
        'client_script_base_url': get_public_app_base_url(),
        'client_script_fallback_url': get_public_script_fallback_url(),
        **kpis 
    }

@bp_tasks.route("/api/pending_tasks")
def api_pending_tasks():
    """API para obtener todas las tareas pendientes para el modal global."""
    try:
        with get_db_connection() as conn:
            tasks = conn.execute("""
                SELECT t.id, t.pc_name, t.descripcion, t.solicitante, t.categoria, 
                       t.estado, t.assigned_to, t.created_at, t.fuero,
                       p.last_user, u.phone
                FROM tasks t
                LEFT JOIN pcs p ON t.pc_name = p.pc_name
                LEFT JOIN ad_users u ON t.solicitante = u.real_name OR t.solicitante = u.username
                WHERE t.estado != 'Hecha'
                ORDER BY t.created_at DESC
            """).fetchall()
            
            # Adjuntar matches y acciones a las tareas pendientes
            tasks = _attach_task_user_matches(tasks, conn)
            tasks = _attach_task_actions_bulk(tasks, conn)
            
            result = []
            for d in tasks:
                if d['created_at']:
                    # Formato legible: "11 Abr, 14:30"
                    d['created_at_fmt'] = d['created_at'].strftime("%d %b, %H:%M")
                
                # Asegurar que las acciones sean diccionarios para jsonify
                if 'acciones' in d and d['acciones']:
                    d['acciones'] = [dict(a) for a in d['acciones']]
                    
                result.append(d)

            return jsonify({
                "status": "success",
                "tasks": result
            })
    except Exception as e:
        print(f"Error en api_pending_tasks: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@bp_tasks.route("/pc/migrate_tasks", methods=["POST"])
def migrate_generic_tasks():
    target_pc = request.form.get("target_pc")
    task_id = request.form.get("migration_task_id")
    if not target_pc: return redirect(url_for("dashboard.pc_detail", pc_name="PC Generica"))
    
    with get_db_connection() as conn:
        if task_id:
            conn.execute("UPDATE tasks SET pc_name = %s WHERE id = %s AND pc_name = 'PC Generica'", (target_pc, task_id))
            audit_msg = f"Se importó la tarea #{task_id} de PC Generica"
            conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                         ('PC Generica', f"Tarea #{task_id} Transferida", "", f"Enviada a {target_pc}", current_username(), "MIGRACION_TAREAS", request.remote_addr))
        else:
            conn.execute("UPDATE tasks SET pc_name = %s WHERE pc_name = 'PC Generica'", (target_pc,))
            audit_msg = "Se importaron TODAS las tareas de PC Generica"
            conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                         ('PC Generica', "Todas las tareas transferidas", "", f"Enviadas a {target_pc}", current_username(), "MIGRACION_TAREAS", request.remote_addr))
        
        conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                     (target_pc, "MIGRACION", "PC Generica", audit_msg, current_username(), "MIGRACION_TAREAS", request.remote_addr))
        conn.commit()
    # Redirigir a la PC destino donde se asigno la tarea
    return redirect(url_for("dashboard.pc_detail", pc_name=target_pc))

@bp_tasks.route("/pc/<pc_name>/tasks", methods=["POST"])
def add_task(pc_name):
    descripcion = request.form.get("descripcion", "").strip()
    solicitante = request.form.get("solicitante", "").strip()
    categoria = request.form.get("categoria", "").strip()
    tipo_actividad = request.form.get("tipo_actividad", "tarea").strip()
    prioridad = request.form.get("prioridad", "1").strip()
    impacto_valor = request.form.get("impacto_valor", "1").strip()
    resumen_impacto = request.form.get("resumen_impacto", "").strip()

    technician = request.form.get("technician", "").strip()

    if not solicitante: solicitante = "No Especificado (Dashboard)"
    if not descripcion: return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))
    if not categoria: categoria = predict_category(descripcion)

    is_done = request.form.get("is_done") == "on"
    solucion = request.form.get("solucion", "").strip()

    estado = "Pendiente"
    assigned_to = None
    completed_by = None
    completed_at = None

    if is_done:
        estado = "Hecha"
        completed_by = technician if technician else current_username()
        completed_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        assigned_to = technician if technician else current_username()
    elif technician:
        assigned_to = technician
        estado = "Asignada"

    with get_db_connection() as conn:
        cursor = conn.execute(
            """INSERT INTO tasks (pc_name, descripcion, solicitante, categoria, tipo_actividad, prioridad, impacto_valor, resumen_impacto, assigned_to, estado, completed_by, completed_at, solucion) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", 
            (pc_name, descripcion, solicitante, categoria, tipo_actividad, prioridad, impacto_valor, resumen_impacto, assigned_to, estado, completed_by, completed_at, solucion)
        )
        task_id = cursor.lastrowid
        conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                     (pc_name, f"Tarea {tipo_actividad.upper()} Creada", "", f"#{task_id}: {descripcion[:30]}..." + (" (Resuelta)" if is_done else ""), current_username(), "GESTION_TAREAS", request.remote_addr))
        conn.commit()

    # Notify technicians
    try:
        from datetime import datetime as _dt
        import locale
        
        phone_info = ""
        with get_db_connection() as conn:
            # Try to match solicitante against real_name or username
            user_row = conn.execute("SELECT phone FROM ad_users WHERE real_name = %s OR username = %s LIMIT 1", (solicitante, solicitante)).fetchone()
            if user_row and user_row["phone"]:
                phone_info = f"📞 *Teléfono:* {user_row['phone']}\n"

        
        # Try to set locale for Spanish dates, fallback if not available
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            try: locale.setlocale(locale.LC_TIME, '') # Try default
            except: pass
            
        fecha_str = _dt.now().strftime("%A %d de %B de %Y").capitalize()
        
        cuerpo = f"📅 *Fecha:* {fecha_str}\n"
        cuerpo += f"🖥️ *PC/Equipo:* {pc_name}\n"
        cuerpo += f"👤 *Solicitante:* {solicitante}\n"
        cuerpo += phone_info
        cuerpo += f"🏷️ *Categoría:* {categoria}\n"
        cuerpo += f"📝 *Descripción:* {descripcion}\n"

        notify_all_technicians(
            title="🚨 Nueva Tarea (PC)",
            body=cuerpo,
            url="/tecnicos"
        )
    except Exception as e:
        print(f"Error notifying: {e}")

    return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))

@bp_tasks.route("/technicians/add", methods=["POST"])
@superuser_required
def add_technician():
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db_connection() as conn:
                conn.execute("INSERT INTO technicians (name) VALUES (%s)", (name,))
                conn.commit()
        except pymysql.err.IntegrityError: pass
    return redirect(url_for("dashboard.dashboard"))

@bp_tasks.route("/technicians/delete/<int:tech_id>", methods=["POST"])
@superuser_required
def delete_technician(tech_id):
    with get_db_connection() as conn:
        conn.execute("DELETE FROM technicians WHERE id = %s", (tech_id,))
        conn.commit()
    return redirect(url_for("dashboard.dashboard"))

@bp_tasks.route("/tasks/<int:task_id>/done", methods=["POST"])
def mark_task_done(task_id):
    technician = request.form.get("technician_name")
    if not technician:
        technician = request.form.get("technician_name_hidden")
    technician = (technician or "").strip()
    selected_pc_name = (request.form.get("pc_name") or "").strip()
    solucion = (request.form.get("solucion") or "").strip()
    if not _is_assignable_technician(technician):
        return redirect(request.referrer or url_for("dashboard.dashboard"))
    with get_db_connection() as conn:
        row = conn.execute("SELECT pc_name FROM tasks WHERE id = %s", (task_id,)).fetchone()
        if row:
            pc_name = row["pc_name"]
            if selected_pc_name:
                pc_name = selected_pc_name
            conn.execute(
                "UPDATE tasks SET estado = 'Hecha', completed_by = %s, completed_at = %s, assigned_to = COALESCE(NULLIF(assigned_to, ''), %s), pc_name = %s, solucion = %s WHERE id = %s",
                (technician, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), technician, pc_name, solucion, task_id)
            )
            conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                         (pc_name, f"Tarea #{task_id} Completada", "Pendiente", f"Por {technician}", current_username(), "GESTION_TAREAS", request.remote_addr))
            conn.commit()
        else: pc_name = ""
    ref = request.referrer
    if ref and request.host in ref and '/pc/' not in ref:
        if '#' in ref: ref = ref.split('#')[0]
        modal_hash = f"#auxModal_{pc_name.replace(' ', '_').replace('-', '_')}" if pc_name else ""
        return redirect(f"{ref}{modal_hash}")
    if not pc_name: return redirect(url_for("dashboard.dashboard"))
    return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))

@bp_tasks.route("/tasks/<int:task_id>/delete", methods=["POST"])
def delete_task(task_id):
    with get_db_connection() as conn:
        row = conn.execute("SELECT pc_name, descripcion FROM tasks WHERE id = %s", (task_id,)).fetchone()
        if row:
            conn.execute("DELETE FROM tasks WHERE id = %s", (task_id,))
            pc_name = row["pc_name"]
            descripcion = row.get("descripcion", "Sin descripción")
            conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                         (pc_name, "Tarea Eliminada", f"#{task_id}", f"Desc: {descripcion[:50]}", current_username(), "GESTION_TAREAS", request.remote_addr))
            conn.commit()
        else: pc_name = ""
    ref = request.referrer
    if ref and request.host in ref and '/pc/' not in ref:
        if '#' in ref: ref = ref.split('#')[0]
        modal_hash = f"#auxModal_{pc_name.replace(' ', '_').replace('-', '_')}" if pc_name else ""
        return redirect(f"{ref}{modal_hash}")
    if not pc_name: return redirect(url_for("dashboard.dashboard"))
    return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))

@bp_tasks.route("/create_loose_task", methods=["POST"])
def create_loose_task():
    descripcion = request.form.get("descripcion")
    solicitante = request.form.get("solicitante")
    categoria = request.form.get("categoria")
    technician = request.form.get("technician")
    fuero = request.form.get("fuero")

    tipo_actividad = request.form.get("tipo_actividad", "tarea").strip()
    prioridad = request.form.get("prioridad", "1").strip()
    impacto_valor = request.form.get("impacto_valor", "1").strip()
    resumen_impacto = request.form.get("resumen_impacto", "").strip()

    if not descripcion or not solicitante: return "Faltan datos", 400
    if not categoria: categoria = predict_category(descripcion)
    
    is_done = request.form.get("is_done") == "on"
    solucion = request.form.get("solucion", "").strip()

    estado = "Pendiente"
    assigned_to = None
    completed_by = None
    completed_at = None

    if is_done:
        estado = "Hecha"
        completed_by = technician if technician else current_username()
        completed_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        assigned_to = technician if technician else current_username()
    elif technician:
        assigned_to = technician
        estado = "Asignada"

    with get_db_connection() as conn:
        cursor = conn.execute(
            """INSERT INTO tasks (descripcion, solicitante, estado, created_at, categoria, assigned_to, fuero, pc_name, tipo_actividad, prioridad, impacto_valor, resumen_impacto, completed_by, completed_at, solucion) 
               VALUES (%s, %s, %s, NOW(), %s, %s, %s, 'PC Generica', %s, %s, %s, %s, %s, %s, %s)""",
            (descripcion, solicitante, estado, categoria, assigned_to, fuero, tipo_actividad, prioridad, impacto_valor, resumen_impacto, completed_by, completed_at, solucion)
        )
        conn.commit()

    # Notify technicians
    try:
        from datetime import datetime as _dt
        import locale
        
        phone_info = ""
        with get_db_connection() as conn:
            user_row = conn.execute("SELECT phone FROM ad_users WHERE real_name = %s OR username = %s LIMIT 1", (solicitante, solicitante)).fetchone()
            if user_row and user_row["phone"]:
                phone_info = f"📞 *Teléfono:* {user_row['phone']}\n"

        
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            try: locale.setlocale(locale.LC_TIME, '')
            except: pass
            
        fecha_str = _dt.now().strftime("%A %d de %B de %Y").capitalize()
        
        cuerpo = f"📅 *Fecha:* {fecha_str}\n"
        cuerpo += f"âš–ï¸ *Fuero/Ãrea:* {fuero if fuero else 'No especificado'}\n"
        cuerpo += f"👤 *Solicitante:* {solicitante}\n"
        cuerpo += phone_info
        cuerpo += f"🏷️ *Categoría:* {categoria}\n"
        if assigned_to:
            cuerpo += f"ðŸ‘¨â€ðŸ”§ *Asignada a:* {assigned_to}\n"
        cuerpo += f"📝 *Descripción:* {descripcion}\n"

        notify_all_technicians(
            title="🚨 Nueva Tarea Suelta",
            body=cuerpo,
            url="/tecnicos"
        )
    except Exception as e:
        print(f"Error notifying: {e}")

    return redirect(url_for("dashboard.dashboard"))

@bp_tasks.route("/tasks/assign", methods=["POST"])
def assign_task():
    task_id = request.form.get("task_id")
    pc_name = request.form.get("pc_name", "").strip()
    technician = request.form.get("technician", "").strip()
    if technician and not _is_assignable_technician(technician):
        technician = ""

    if task_id and pc_name:
        with get_db_connection() as conn:
            # Debugging logs (se pueden ver en los logs del servidor)
            print(f"[DEBUG] Asignando tarea {task_id} a PC {pc_name}")
            
            t = conn.execute("SELECT 1 FROM pcs WHERE pc_name = %s", (pc_name,)).fetchone()
            if t:
                # Obtenemos la tarea para ver su estado actual y de donde venía
                row = conn.execute("SELECT pc_name, estado FROM tasks WHERE id = %s", (task_id,)).fetchone()
                if not row:
                    print(f"[DEBUG] No se encontró la tarea {task_id}")
                    return redirect(url_for("dashboard.dashboard"))
                
                old_pc = row['pc_name'] or 'PC Generica'
                current_estado = row['estado']
                
                # Si la tarea ya estaba 'Hecha', la dejamos como 'Hecha' al moverla de PC
                if current_estado == 'Hecha':
                    estado = 'Hecha'
                else:
                    estado = "Asignada" if technician else "Pendiente"
                
                assigned_to = technician if technician else None

                conn.execute(
                    "UPDATE tasks SET pc_name = %s, assigned_to = COALESCE(%s, assigned_to), estado = %s WHERE id = %s",
                    (pc_name, assigned_to, estado, task_id)
                )
                
                # Log on both the new and old PC if they differ
                tech_log = f" (Técnico: {technician})" if technician else ""
                conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                             (pc_name, f"Tarea #{task_id} Asignada aquí", old_pc, f"Asignada a {pc_name}{tech_log}", current_username(), "GESTION_TAREAS", request.remote_addr))
                if old_pc and old_pc != pc_name:
                    conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                                 (old_pc, f"Tarea #{task_id} Transferida", pc_name, f"Transferida a {pc_name}", current_username(), "GESTION_TAREAS", request.remote_addr))
                conn.commit()
                print(f"[DEBUG] Tarea {task_id} actualizada correctamente a {pc_name}")
            else:
                print(f"[DEBUG] No se encontró la PC {pc_name} en la base de datos")
    return redirect(url_for("dashboard.dashboard"))

@bp_tasks.route("/tasks/reassign", methods=["POST"])
def reassign_task():
    task_id = request.form.get("task_id")
    technician = (request.form.get("technician") or "").strip()
    pc_name = request.form.get("pc_name")
    
    if task_id and technician and _is_assignable_technician(technician):
        with get_db_connection() as conn:
            conn.execute("UPDATE tasks SET assigned_to = %s, estado = 'Asignada' WHERE id = %s", (technician, task_id))
            
            if not pc_name:
                row = conn.execute("SELECT pc_name FROM tasks WHERE id = %s", (task_id,)).fetchone()
                if row:
                    pc_name = row["pc_name"]
                    
            conn.commit()
    
    if pc_name:
        ref = request.referrer
        if ref and request.host in ref and '/pc/' not in ref:
            if '#' in ref: ref = ref.split('#')[0]
            modal_hash = f"#auxModal_{pc_name.replace(' ', '_').replace('-', '_')}"
            return redirect(f"{ref}{modal_hash}")
        return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))
    return redirect(url_for("dashboard.dashboard"))

@bp_tasks.route("/api/audit/<pc_name>/add", methods=["POST"])
def add_manual_audit(pc_name):
    campo = request.form.get("campo", "").strip()
    valor_anterior = request.form.get("valor_anterior", "").strip()
    valor_nuevo = request.form.get("valor_nuevo", "").strip()
    if not campo or not valor_nuevo: return jsonify({"status": "error", "message": "Faltan datos"}), 400
    try:
        with get_db_connection() as conn:
            conn.execute("INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                         (pc_name, campo, valor_anterior, valor_nuevo, current_username(), "AUDITORIA_MANUAL", request.remote_addr))
            conn.commit()
        return redirect(url_for("dashboard.pc_detail", pc_name=pc_name))
    except Exception as e:
        return f"Error agregando historial manual: {e}", 500

@bp_tasks.route("/api/report/preview", methods=["GET"])
def report_preview():
    """API para vista previa de tareas antes de descargar el reporte."""
    fecha_filtro = request.args.get("fecha", dt.now().strftime("%Y-%m-%d")).strip()
    pc_name = request.args.get("pc", "").strip()

    with get_db_connection() as conn:
        # Tareas completadas
        sql_hechas = "SELECT t.pc_name, t.descripcion, t.solicitante, t.completed_at, t.completed_by, p.last_user FROM tasks t LEFT JOIN pcs p ON t.pc_name = p.pc_name WHERE t.estado = 'Hecha' AND DATE(t.completed_at) = %s"
        params = [fecha_filtro]
        if pc_name:
            sql_hechas += " AND t.pc_name = %s"
            params.append(pc_name)
        sql_hechas += " ORDER BY t.completed_at DESC"
        hechas = conn.execute(sql_hechas, params).fetchall()

        # Tareas pendientes del dia
        sql_pend = "SELECT t.pc_name, t.descripcion, t.solicitante, t.created_at, t.estado, t.assigned_to FROM tasks t WHERE t.estado != 'Hecha' AND DATE(t.created_at) = %s"
        params_pend = [fecha_filtro]
        if pc_name:
            sql_pend += " AND t.pc_name = %s"
            params_pend.append(pc_name)
        sql_pend += " ORDER BY t.created_at DESC"
        pendientes = conn.execute(sql_pend, params_pend).fetchall()

    return jsonify({
        "fecha": fecha_filtro,
        "pc_filter": pc_name or None,
        "hechas": [{
            "pc_name": r["pc_name"], "descripcion": r["descripcion"],
            "solicitante": r["solicitante"] or "",
            "completed_at": r["completed_at"].strftime("%Y-%m-%d %H:%M:%S") if r["completed_at"] else "",
            "completed_by": r["completed_by"] or "",
        } for r in hechas],
        "pendientes": [{
            "pc_name": r["pc_name"], "descripcion": r["descripcion"],
            "solicitante": r["solicitante"] or "",
            "created_at": r["created_at"].strftime("%Y-%m-%d %H:%M:%S") if r["created_at"] else "",
            "estado": r["estado"] or "",
            "assigned_to": r["assigned_to"] or "Sin asignar",
        } for r in pendientes]
    })

@bp_tasks.route("/report/tasks_completed", methods=["GET", "POST"])
def report_tasks_completed():
    pc_name = request.args.get("pc", "").strip()
    if request.method == "GET": return render_template("report_tasks.html", pc_name=pc_name)

    fecha_filtro = request.form.get("fecha", "").strip() or dt.now().strftime("%Y-%m-%d")
    pc_name = request.form.get("pc_name", "").strip() or pc_name

    with get_db_connection() as conn:
        base_sql = "SELECT pc_name, descripcion, solicitante, created_at, estado, completed_by, completed_at FROM tasks WHERE estado = 'Hecha' AND DATE(completed_at) = %s"
        params = [fecha_filtro]
        if pc_name:
            base_sql += " AND pc_name = %s"
            params.append(pc_name)
        base_sql += " ORDER BY completed_at DESC"
        tareas = conn.execute(base_sql, params).fetchall()

    # --- Excel con formato mejorado ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas Completadas"

    # Estilo de headers
    header_fill = PatternFill("solid", fgColor="0F4C75")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["PC", "Descripción", "Solicitante", "Fecha Creación", "Estado", "Realizado Por", "Fecha Cierre"]
    ws.append(headers)
    for col_idx, col_letter in enumerate([get_column_letter(i+1) for i in range(len(headers))], 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[1].height = 22

    # Estilos alternados para filas
    fill_even = PatternFill("solid", fgColor="EFF6FF")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    data_align_wrap = Alignment(wrap_text=True, vertical="top")

    for row_idx, t in enumerate(tareas, 2):
        row_data = [t["pc_name"], t["descripcion"], t["solicitante"] or "", t["created_at"], t["estado"], t["completed_by"] or "", t["completed_at"] or ""]
        ws.append(row_data)
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = data_font
            cell.border = cell_border
            cell.alignment = data_align_wrap

    # Auto-ancho de columnas
    col_widths = [20, 40, 20, 20, 12, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fila de totales
    ws.append([])
    total_row = ws.max_row + 1
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.value = f"Total tareas: {len(tareas)}"
    total_cell.font = Font(bold=True, name="Calibri", size=10, color="0F4C75")

    # Freeze header
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_pc_sufijo = f"_{pc_name}" if pc_name else ""
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Tareas_Completadas{nombre_pc_sufijo}_{fecha_filtro}.xlsx")

@bp_tasks.route("/report/tasks_completed_pdf", methods=["POST"])
def report_tasks_completed_pdf():
    fecha_filtro_str = request.form.get("fecha", "").strip() or dt.now().strftime("%Y-%m-%d")
    pc_name = request.form.get("pc_name", "").strip()
    fecha_dt = datetime.datetime.strptime(fecha_filtro_str, "%Y-%m-%d")
    fecha_display = format_date_es(fecha_dt)

    with get_db_connection() as conn:
        base_sql = "SELECT t.pc_name, t.descripcion, t.solicitante, t.created_at, t.estado, t.completed_by, t.completed_at, p.last_user FROM tasks t LEFT JOIN pcs p ON t.pc_name = p.pc_name WHERE t.estado = 'Hecha' AND DATE(t.completed_at) = %s"
        params = [fecha_filtro_str]
        if pc_name:
            base_sql += " AND t.pc_name = %s"
            params.append(pc_name)
        base_sql += " ORDER BY t.completed_at DESC"
        tareas_hechas = conn.execute(base_sql, params).fetchall()

        sql_pendientes = "SELECT t.pc_name, t.descripcion, t.solicitante, t.created_at, t.estado, t.assigned_to, p.last_user FROM tasks t LEFT JOIN pcs p ON t.pc_name = p.pc_name WHERE t.estado != 'Hecha' AND DATE(t.created_at) = %s"
        params_pend = [fecha_filtro_str]
        if pc_name:
            sql_pendientes += " AND t.pc_name = %s"
            params_pend.append(pc_name)
        sql_pendientes += " ORDER BY t.created_at DESC"
        tareas_pendientes = conn.execute(sql_pendientes, params_pend).fetchall()

    def draw_tasks_table(pdf, tasks, widths, headers, colors):
        # Header
        pdf.set_font("Arial", "B", 9)
        pdf.set_fill_color(*colors['header_bg'])
        pdf.set_text_color(255)
        for i, h in enumerate(headers):
            pdf.cell(widths[i], 8, h, 0, 0, 'C', fill=True)
        pdf.ln()

        if not tasks:
            pdf.set_font("Arial", "I", 9)
            pdf.set_text_color(100)
            pdf.cell(sum(widths), 10, "No hay registros para hoy.", 'B', 1, 'C')
            return

        # Data Rows
        pdf.set_font("Arial", "", 8)
        pdf.set_text_color(0)
        row_idx = 0
        for t in tasks:
            # Prepare data
            if 'completed_by' in t.keys(): # Hechas
                raw_user = t["last_user"] or "N/A"
                user_display = raw_user.split("\\\\")[-1] if "\\\\" in raw_user else raw_user
                cols = [
                    str(t["pc_name"]),
                    str(user_display),
                    str(t["descripcion"] or ""),
                    str(t["solicitante"] or ""),
                    str(t["completed_by"] or ""),
                    format_datetime_es(t["created_at"])
                ]
            else: # Pendientes
                cols = [
                    str(t["pc_name"] or "N/A"),
                    str(t["descripcion"] or ""),
                    str(t["solicitante"] or ""),
                    str(t["assigned_to"] or "Sin Asignar"),
                    format_datetime_es(t["created_at"])
                ]

            # Calculate height
            row_height = 5
            max_lines = 1
            split_cols = []
            for i, text in enumerate(cols):
                lines = pdf.multi_cell(widths[i], row_height, text, split_only=True)
                max_lines = max(max_lines, len(lines))
                split_cols.append(lines)
            
            h_row = max_lines * row_height + 2 # + padding

            # Page break
            if (pdf.get_y() + h_row) > 275:
                pdf.add_page()
                # Redibujar cabecera
                pdf.set_font("Arial", "B", 9)
                pdf.set_fill_color(*colors['header_bg'])
                pdf.set_text_color(255)
                for i, h in enumerate(headers): pdf.cell(widths[i], 8, h, 0, 0, 'C', fill=True)
                pdf.ln()
                pdf.set_font("Arial", "", 8)
                pdf.set_text_color(0)

            # Draw
            x_row, y_row = pdf.get_x(), pdf.get_y()
            
            # Row Background (Striped)
            if row_idx % 2 == 1:
                pdf.set_fill_color(248, 249, 250)
                pdf.rect(x_row, y_row, sum(widths), h_row, "F")
            
            pdf.set_text_color(0)
            # Draw each cell
            for i, lines in enumerate(split_cols):
                pdf.set_xy(x_row + sum(widths[:i]), y_row + 1)
                cell_text = "\n".join(lines)
                pdf.multi_cell(widths[i], row_height, cell_text, 0, 'L' if i != len(widths)-1 else 'C')
            
            # Bottom line (Border)
            pdf.set_draw_color(230, 230, 230)
            pdf.line(x_row, y_row + h_row, x_row + sum(widths), y_row + h_row)
            
            pdf.set_xy(x_row, y_row + h_row)
            row_idx += 1

    # --- MAIN PDF CONTENT ---
    pdf = PDFReport(title="Reporte de Tareas - Inventario GOLD")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Reporte del día: {fecha_display}", 0, 1)
    pdf.ln(2)

    # Tabla Hechas
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(25, 135, 84)
    pdf.cell(0, 8, f"Tareas Realizadas ({len(tareas_hechas)})", 0, 1)
    pdf.ln(1)
    
    widths_hechas = [32, 25, 50, 22, 22, 39]
    headers_hechas = ["PC", "Usuario", "Descripción", "Solic.", "Técnico", "Creada el"]
    draw_tasks_table(pdf, tareas_hechas, widths_hechas, headers_hechas, {'header_bg': (25, 135, 84)})

    pdf.ln(5)

    # Tabla Pendientes
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(220, 53, 69)
    pdf.cell(0, 8, f"Tareas Pendientes / Generadas Hoy ({len(tareas_pendientes)})", 0, 1)
    pdf.ln(1)

    widths_pend = [30, 70, 30, 30, 30]
    headers_pend = ["PC", "Descripción", "Solicitante", "Asignado a", "Creada el"]
    draw_tasks_table(pdf, tareas_pendientes, widths_pend, headers_pend, {'header_bg': (180, 50, 60)})

    output = BytesIO()
    pdf_bytes = pdf.output()
    output.write(pdf_bytes)
    output.seek(0)
    
    nombre_pc_sufijo = f"_{pc_name}" if pc_name else ""
    return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=f"Reporte_Tareas{nombre_pc_sufijo}_{fecha_filtro_str}.pdf")


def _attach_task_actions_bulk(tasks, conn):
    if not tasks: return tasks
    task_ids = [t['id'] for t in tasks]
    placeholders = ",".join(["%s"] * len(task_ids))
    actions = conn.execute(f"SELECT id, task_id, user_name, action_text, created_at FROM task_actions WHERE task_id IN ({placeholders}) ORDER BY created_at ASC", tuple(task_ids)).fetchall()
    
    actions_map = {tid: [] for tid in task_ids}
    for a in actions:
        a_dict = dict(a)
        if a_dict["created_at"]:
            a_dict["created_at_fmt"] = a_dict["created_at"].strftime("%d/%m %H:%M")
            a_dict["created_at"] = a_dict["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        actions_map[a['task_id']].append(a_dict)
        
    for t in tasks:
        t['acciones'] = actions_map.get(t['id'], [])
    return tasks

@bp_tasks.route("/visor")
def visor():
    """Vista de 'visor' para mostrar trabajos con filtros."""
    try:
        hoy_str = dt.now().strftime("%Y-%m-%d")
        fecha_filtro = request.args.get("fecha", hoy_str)
        pc_filtro = request.args.get("pc", "").strip()
        tech_filtro = request.args.get("technician", "").strip()
        is_filtered = bool(pc_filtro or tech_filtro or (request.args.get("fecha") and request.args.get("fecha") != hoy_str))

        with get_db_connection() as conn:
            technicians = list_technician_users()
            
            if is_filtered:
                base_sql = "SELECT t.*, p.last_user FROM tasks t LEFT JOIN pcs p ON t.pc_name = p.pc_name WHERE 1=1"
                params = []
                if request.args.get("fecha"):
                    base_sql += " AND (DATE(t.created_at) = %s OR (t.estado = 'Hecha' AND DATE(t.completed_at) = %s))"
                    params.extend([fecha_filtro, fecha_filtro])
                if pc_filtro:
                    base_sql += " AND t.pc_name LIKE %s"
                    params.append(f"%{pc_filtro}%")
                if tech_filtro:
                    base_sql += " AND (t.assigned_to = %s OR t.completed_by = %s)"
                    params.extend([tech_filtro, tech_filtro])
                
                base_sql += " ORDER BY t.created_at DESC LIMIT 200"
                tareas = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute(base_sql, params).fetchall(), conn), conn)
                
                return render_template("visor_tareas.html", 
                                       tareas_hoy=[_decorate_visor_task(t) for t in tareas], 
                                       tareas_anteriores=[],
                                       tareas_pendientes=[],
                                       technicians=technicians,
                                       hoy=fecha_filtro,
                                       pc_filtro=pc_filtro,
                                       tech_filtro=tech_filtro,
                                       is_filtered=True)
            else:
                tareas_hoy = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE DATE(t.created_at) = CURDATE() OR (t.estado = 'Hecha' AND DATE(t.completed_at) = CURDATE())
                    ORDER BY t.created_at DESC
                """).fetchall(), conn), conn)
                tareas_anteriores = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE DATE(t.created_at) < CURDATE() AND DATE(t.created_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                    ORDER BY t.created_at DESC LIMIT 50
                """).fetchall(), conn), conn)
                tareas_pendientes = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE t.estado != 'Hecha'
                    ORDER BY t.created_at DESC
                """).fetchall(), conn), conn)
                return render_template("visor_tareas.html", 
                                       tareas_hoy=[_decorate_visor_task(t) for t in tareas_hoy], 
                                       tareas_anteriores=[_decorate_visor_task(t) for t in tareas_anteriores],
                                       tareas_pendientes=[_decorate_visor_task(t) for t in tareas_pendientes],
                                       technicians=technicians,
                                       hoy=fecha_filtro,
                                       is_filtered=False)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Error cargando el visor: {e}", 500

@bp_tasks.route("/api/visor/data")
def api_visor_data():
    """API para actualización en tiempo real del visor con soporte de filtros."""
    try:
        hoy_str = dt.now().strftime("%Y-%m-%d")
        fecha_filtro = request.args.get("fecha")
        pc_filtro = request.args.get("pc", "").strip()
        tech_filtro = request.args.get("technician", "").strip()
        is_filtered = bool(pc_filtro or tech_filtro or (fecha_filtro and fecha_filtro != hoy_str))

        with get_db_connection() as conn:
            if is_filtered:
                base_sql = "SELECT t.*, p.last_user FROM tasks t LEFT JOIN pcs p ON t.pc_name = p.pc_name WHERE 1=1"
                params = []
                if fecha_filtro:
                    base_sql += " AND (DATE(t.created_at) = %s OR (t.estado = 'Hecha' AND DATE(t.completed_at) = %s))"
                    params.extend([fecha_filtro, fecha_filtro])
                if pc_filtro:
                    base_sql += " AND t.pc_name LIKE %s"
                    params.append(f"%{pc_filtro}%")
                if tech_filtro:
                    base_sql += " AND (t.assigned_to = %s OR t.completed_by = %s)"
                    params.extend([tech_filtro, tech_filtro])
                
                base_sql += " ORDER BY t.created_at DESC LIMIT 200"
                rows = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute(base_sql, params).fetchall(), conn), conn)
                
                return jsonify({
                    "status": "success", 
                    "tasks_hoy": [_decorate_visor_task(t) for t in rows],
                    "tasks_anteriores": [],
                    "tasks_pendientes": []
                })
            else:
                # Tareas de hoy
                tareas_hoy_rows = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user 
                    FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE DATE(t.created_at) = CURDATE() OR (t.estado = 'Hecha' AND DATE(t.completed_at) = CURDATE())
                    ORDER BY t.created_at DESC
                """).fetchall(), conn), conn)
                
                # Tareas anteriores (pendientes o recientes)
                tareas_anteriores_rows = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user 
                    FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE DATE(t.created_at) < CURDATE() 
                    AND DATE(t.created_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                    ORDER BY t.created_at DESC
                    LIMIT 50
                """).fetchall(), conn), conn)

                # Tareas pendientes
                tareas_pendientes_rows = _attach_task_actions_bulk(_attach_task_user_matches(conn.execute("""
                    SELECT t.*, p.last_user 
                    FROM tasks t 
                    LEFT JOIN pcs p ON t.pc_name = p.pc_name 
                    WHERE t.estado != 'Hecha'
                    ORDER BY t.created_at DESC
                """).fetchall(), conn), conn)

                return jsonify({
                    "status": "success", 
                    "tasks_hoy": [_decorate_visor_task(t) for t in tareas_hoy_rows],
                    "tasks_anteriores": [_decorate_visor_task(t) for t in tareas_anteriores_rows],
                    "tasks_pendientes": [_decorate_visor_task(t) for t in tareas_pendientes_rows]
                })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@bp_tasks.route("/api/tasks/<int:task_id>/actions", methods=["GET"])
def get_task_actions(task_id):
    try:
        with get_db_connection() as conn:
            actions = conn.execute(
                "SELECT id, user_name, action_text, created_at FROM task_actions WHERE task_id = %s ORDER BY created_at ASC",
                (task_id,)
            ).fetchall()
            
            result = []
            for a in actions:
                a_dict = dict(a)
                if a_dict.get("created_at"):
                    if hasattr(a_dict["created_at"], "strftime"):
                        a_dict["created_at_fmt"] = a_dict["created_at"].strftime("%d/%m/%Y %H:%M:%S")
                        a_dict["created_at"] = a_dict["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        a_dict["created_at_fmt"] = str(a_dict["created_at"])
                        a_dict["created_at"] = str(a_dict["created_at"])
                result.append(a_dict)
                
            return jsonify({"status": "success", "actions": result})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@bp_tasks.route("/api/tasks/<int:task_id>/actions", methods=["POST"])
def add_task_action(task_id):
    if request.is_json:
        action_text = request.json.get("action_text", "").strip()
    else:
        action_text = request.form.get("action_text", "").strip()
        
    if not action_text:
        return jsonify({"status": "error", "message": "El texto de la acción no puede estar vacío"}), 400
        
    try:
        with get_db_connection() as conn:
            from utils.auth import current_technician_identity
            username = current_username() or current_technician_identity() or "Desconocido"
            
            conn.execute(
                "INSERT INTO task_actions (task_id, user_name, action_text) VALUES (%s, %s, %s)",
                (task_id, username, action_text)
            )
            
            # Log it in audit_logs
            pc_row = conn.execute("SELECT pc_name FROM tasks WHERE id = %s", (task_id,)).fetchone()
            if pc_row:
                pc_name = pc_row["pc_name"]
                conn.execute(
                    "INSERT INTO audit_logs (pc_name, field, old_value, new_value, user_name, action_type, ip_address) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (pc_name, f"Acción en Tarea #{task_id}", "", action_text[:100], username, "ACCION_TAREA", request.remote_addr)
                )
            
            conn.commit()
            
            try:
                from services.push_notifications import notify_all_technicians
                notify_all_technicians(
                    title=f"Nueva nota en Tarea #{task_id}",
                    body=f"{username}: {action_text}",
                    url="/tecnicos"
                )
            except Exception as e:
                print(f"Error notifying action: {e}")

            return jsonify({"status": "success", "message": "Acción agregada correctamente"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
