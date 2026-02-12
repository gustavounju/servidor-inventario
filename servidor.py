from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, make_response
import json
import os
import sqlite3
import datetime
import csv
import io
from openpyxl import Workbook
from datetime import datetime as dt
from io import BytesIO
import socket
from classifier import SimpleNaiveBayes, SEED_DATA
from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__)
# Aplicar ProxyFix para que Flask entienda que está detrás de (Nginx) y maneje bien HTTPS/URLs
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
# Inicializar IA
ai_classifier = SimpleNaiveBayes()

app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

DB_FILE = "inventario.db"
LOG_FOLDER = "logs"
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'pdf'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ----------------- Utilidades DB -----------------

def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Crea tabla pcs y tabla tasks con todas las columnas necesarias."""
    print(f"Inicializando base de datos en '{DB_FILE}'...")
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS pcs (
                pc_name TEXT PRIMARY KEY,
                os_name TEXT,
                processor TEXT,
                ram_gb REAL,
                ip_address TEXT,
                last_user TEXT,
                last_report TEXT,
                ram_detalles TEXT,
                disk_models TEXT,
                disk_speeds_rpm TEXT,
                motherboard_model TEXT,
                monitors TEXT,
                printer_model TEXT,
                printer_port TEXT,
                ping_ms TEXT,
                ping_loss_pct TEXT,
                alerta_ram_baja INTEGER DEFAULT 0,
                alerta_sin_impresora INTEGER DEFAULT 0,
                alerta_impresora_red INTEGER DEFAULT 0,
                is_active TEXT DEFAULT 'True',
                full_json_data TEXT
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pc_name TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now', '-3 hours')),
                descripcion TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'Pendiente',
                solicitante TEXT,
                FOREIGN KEY (pc_name) REFERENCES pcs(pc_name)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS technicians (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
            """
        )

        conn.commit()
    print("Base de datos lista y estructura verificada.")

def migrate_db_v2():
    """Migra la tabla tasks para permitir pc_name NULL y agregar solicitante."""
    print("Verificando migración de DB v2...")
    with get_db_connection() as conn:
        # Verificar si ya existe la columna solicitante
        cursor = conn.execute("PRAGMA table_info(tasks)")
        columns = [row["name"] for row in cursor.fetchall()]
        if "solicitante" in columns:
            print("DB ya está en v2.")
            return

        print("Migrando DB a v2...")
        try:
            # 1. Renombrar tabla vieja
            conn.execute("ALTER TABLE tasks RENAME TO tasks_old")

            # 2. Crear tabla nueva
            conn.execute(
                """
                CREATE TABLE tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    pc_name TEXT,
                    created_at TEXT NOT NULL DEFAULT (datetime('now', '-3 hours')),
                    descripcion TEXT NOT NULL,
                    estado TEXT NOT NULL DEFAULT 'Pendiente',
                    solicitante TEXT,
                    FOREIGN KEY (pc_name) REFERENCES pcs(pc_name)
                )
                """
            )

            # 3. Copiar datos
            conn.execute(
                """
                INSERT INTO tasks (id, pc_name, created_at, descripcion, estado)
                SELECT id, pc_name, created_at, descripcion, estado FROM tasks_old
                """
            )

            # 4. Borrar vieja
            conn.execute("DROP TABLE tasks_old")
            conn.commit()
            print("Migración v2 completada con éxito.")
        except Exception as e:
            print(f"Error en migración v2: {e}")
            conn.rollback()

def migrate_db_v3():
    """Migra BD a v3: tabla technicians y columnas completed_by/at en tasks."""
    print("Verificando migración de DB v3...")
    with get_db_connection() as conn:
        # Check if technicians table exists
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='technicians'")
        if not cursor.fetchone():
            print("Creando tabla technicians...")
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS technicians (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE
                )
                """
            )
        
        # Check columns in tasks
        cursor = conn.execute("PRAGMA table_info(tasks)")
        columns = [row["name"] for row in cursor.fetchall()]
        
        if "completed_by" not in columns:
            print("Agregando col completed_by a tasks...")
            try:
                conn.execute("ALTER TABLE tasks ADD COLUMN completed_by TEXT")
            except Exception as e: print(f"Error add completed_by: {e}")

        if "completed_at" not in columns:
            print("Agregando col completed_at a tasks...")
            try:
                conn.execute("ALTER TABLE tasks ADD COLUMN completed_at TEXT")
            except Exception as e: print(f"Error add completed_at: {e}")
            
        conn.commit()
    print("Migración v3 verificada.")

def migrate_db_v4():
    """Migra BD a v4: agregar columna categoria a tasks."""
    print("Verificando migración de DB v4...")
    with get_db_connection() as conn:
        cursor = conn.execute("PRAGMA table_info(tasks)")
        columns = [row["name"] for row in cursor.fetchall()]
        
        if "categoria" not in columns:
            print("Agregando col categoria a tasks...")
            try:
                conn.execute("ALTER TABLE tasks ADD COLUMN categoria TEXT")
                conn.commit()
            except Exception as e: print(f"Error add categoria: {e}")
            
    print("Migración v4 verificada.")

def migrate_db_v5():
    """Migra BD a v5: crear tabla audit_logs para historial de cambios."""
    print("Verificando migración de DB v5...")
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pc_name TEXT,
                field TEXT,
                old_value TEXT,
                new_value TEXT,
                changed_at DATETIME DEFAULT (datetime('now', '-3 hours'))
            )
            """
        )
    print("Migración v5 verificada.")

@app.route("/api/mobile/parse_voice", methods=["POST"])
def api_parse_voice():
    """Recibe audio transcrito/texto y usa IA para extraer estructura."""
    data = request.json
    raw_text = data.get("text", "")
    
    if not raw_text:
        return jsonify({"status": "error", "message": "No text provided"})

    try:
        from voice_processor import process_voice_command
        result = process_voice_command(raw_text)
        return jsonify({"status": "success", "data": result})
    except ImportError:
         return jsonify({"status": "error", "message": "Module voice_processor missing"})
    except Exception as e:
        print(f"Error AI Parse: {e}")
        # Fallback a lo básico
        return jsonify({
            "status": "success", 
            "data": {"descripcion": raw_text, "solicitante": ""}
        })

def migrate_db_v6():
    """Migración V6: Agregar columna 'assigned_to' a la tabla 'tasks'."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        # Verificar si la columna ya existe
        cursor.execute("PRAGMA table_info(tasks)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'assigned_to' not in columns:
            print("Aplicando migración V6: Agregando 'assigned_to' a tasks...")
            cursor.execute("ALTER TABLE tasks ADD COLUMN assigned_to TEXT")
            conn.commit()
        else:
            print("Migración V6 verificada.")

def migrate_db_v7():
    """Migración V7: Agregar columna 'fuero' a la tabla 'pcs'."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        # Verificar si la columna ya existe
        cursor.execute("PRAGMA table_info(pcs)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'fuero' not in columns:
            print("Aplicando migración V7: Agregando 'fuero' a pcs...")
            cursor.execute("ALTER TABLE pcs ADD COLUMN fuero TEXT")
            conn.commit()
        else:
            print("Migración V7 verificada.")

def migrate_db_v8():
    """Migración V8: Agregar columna 'fuero' a la tabla 'tasks'."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(tasks)")
        columns = [info[1] for info in cursor.fetchall()]
        if 'fuero' not in columns:
            print("Aplicando migración V8: Agregando 'fuero' a tasks...")
            cursor.execute("ALTER TABLE tasks ADD COLUMN fuero TEXT")
            conn.commit()
        else:
            print("Migración V8 verificada.")

def migrate_db_v9():
    """Migración V9: Agregar columnas de infraestructura de red y ubicación a 'pcs'."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(pcs)")
        columns = [info[1] for info in cursor.fetchall()]
        
        new_columns = {
            "switch_name": "TEXT",
            "switch_port": "TEXT",
            "pachera_name": "TEXT",
            "pachera_port": "TEXT",
            "building": "TEXT",
            "floor": "TEXT"
        }
        
        for col, dtype in new_columns.items():
            if col not in columns:
                print(f"Aplicando migración V9: Agregando '{col}' a pcs...")
                try:
                    cursor.execute(f"ALTER TABLE pcs ADD COLUMN {col} {dtype}")
                except Exception as e:
                    print(f"Error agregando columna {col}: {e}")
        
        conn.commit()
    print("Migración V9 verificada.")

# --- DICCIONARIO DE FUEROS (CONFIGURABLE) ---
# Agregar aquí nuevos códigos. El sistema buscará el prefijo más largo que coincida.
FUERO_MAPPING = {
    "TTSIVVOC": "Tribunal de Trabajo Sala IV",
    "OGL": "Oficina de Gestion Laboral",
    "SISTEMAS": "Dpto. Informatica San Pedro",
    "VGS": "Violencia de Género 5",
    "VG5": "Violencia de Género 5",
    "SIVL": "Sala IV Laboral",
    "TJO1": "Tribunal de Juicio",
    "TJ01": "Tribunal de Juicio",
    "TJ": "Tribunal de Juicio",
    "CGESE": "Cámara Gesell"
}

FUERO_COLORS = {
    "Tribunal de Trabajo Sala IV": "#0d6efd",    # Blue
    "Oficina de Gestion Laboral": "#198754",     # Green
    "Dpto. Informatica San Pedro": "#212529",    # Dark
    "Violencia de Género 5": "#d63384",          # Pink
    "Sala IV Laboral": "#fd7e14",                # Orange
    "Tribunal de Juicio": "#6610f2",             # Indigo
    "Cámara Gesell": "#20c997"                   # Teal
}

def detect_fuero(pc_name):
    """Detecta el fuero basado en el prefijo del nombre de la PC."""
    if not pc_name:
        return "Desconocido"
    
    pc_upper = pc_name.upper()
    
    # Ordenar por longitud de clave descendente para coincidir prefijos más largos primero
    # Ej: Si existe 'TT' y 'TTVOC', y la PC es 'TTVOC01', coincidirá primero 'TTVOC'.
    for prefix in sorted(FUERO_MAPPING.keys(), key=len, reverse=True):
        if pc_upper.startswith(prefix):
            return FUERO_MAPPING[prefix]
            
    return "Desconocido"

def train_ai_model():
    """Entrena la IA con datos semilla + datos históricos de la DB."""
    try:
        # 1. Entrenar con Semilla (Base Knowledge)
        ai_classifier.train(SEED_DATA)
        print(f"IA: Entrenada con {len(SEED_DATA)} ejemplos semilla.")

        # 2. Entrenar con Datos del Usuario (Incremental Learning)
        conn = get_db_connection()
        tasks = conn.execute("SELECT descripcion, categoria FROM tasks WHERE categoria IS NOT NULL AND categoria != ''").fetchall()
        conn.close()

        user_data = [(t['descripcion'], t['categoria']) for t in tasks]
        if user_data:
            ai_classifier.train(user_data)
            print(f"IA: Refinada con {len(user_data)} tareas históricas.")
        
    except Exception as e:
        print(f"Error entrenando IA: {e}")

# Entrenar al iniciar (después de definir la clase y migraciones)
with app.app_context():
    init_db()
    migrate_db_v2()
    migrate_db_v3()
    migrate_db_v4()
    migrate_db_v5()
    migrate_db_v6()
    migrate_db_v7()
    migrate_db_v8()
    migrate_db_v9()
    
    def ensure_generic_pc():
        """Asegura que exista una PC genérica para asignar tareas a hardware no inventariado."""
        try:
            with get_db_connection() as conn:
                exists = conn.execute("SELECT 1 FROM pcs WHERE pc_name = 'PC Generica'").fetchone()
                if not exists:
                    print("Creando 'PC Generica'...")
                    conn.execute(
                        """
                        INSERT INTO pcs (pc_name, os_name, is_active) 
                        VALUES ('PC Generica', 'Virtual/Pendiente', 'True')
                        """
                    )
                    conn.commit()
        except Exception as e:
            print(f"Error creando PC Generica: {e}")

    ensure_generic_pc()
    train_ai_model()




def predict_category(descripcion):
    """Clasifica la tarea usando el modelo Naive Bayes."""
    if not descripcion:
        return "General"
    
    # Predicción IA
    prediction = ai_classifier.predict(descripcion)
    
    # (Opcional: Logs para depurar qué está pensando)
    # print(f"DEBUG IA: '{descripcion}' -> {prediction}")
    
    return prediction



# ----------------- Healthcheck -----------------

# ----------------- Backup Remoto -----------------
@app.route("/download_db")
def download_db():
    try:
        if not os.path.exists(DB_FILE):
            return "Base de datos no encontrada", 404
            
        # Nombre del archivo con fecha
        filename = f"inventario_backup_{dt.now().strftime('%Y%m%d_%H%M')}.db"
        
        return send_file(
            DB_FILE,
            as_attachment=True,
            download_name=filename,
            mimetype="application/x-sqlite3"
        )
    except Exception as e:
        return f"Error generando backup: {e}", 500


# ----------------- Healthcheck -----------------

@app.route("/health", methods=["GET"])
def health():
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.execute("SELECT 1")
        conn.close()
        return {"status": "ok", "db_ok": True}, 200
    except Exception:
        return {"status": "error", "db_ok": False}, 500


# Crear carpeta de logs
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# ----------------- Rutas Web -----------------

@app.route("/cementerio")
def view_cementerio():
    return redirect(url_for("dashboard", estado="False"))

@app.route("/graficos")
def view_graphics():
    """Nueva página dedicada a KPIs y Gráficos."""
    try:
        with get_db_connection() as conn:
            # 1. Total Activas (Excluyendo PC Generica e Infraestructura)
            kpi_total_activas = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND pc_name NOT IN ('PC Generica', 'Infraestructura')"
            ).fetchone()["c"]

            # 2. Total Cementerio
            kpi_total_graveyard = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'False'"
            ).fetchone()["c"]

            # 3. Alertas (RAM, Sin Impresora, Impresora Red)
            kpi_alerta_ram = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_ram_baja = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            kpi_sin_impresora = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_sin_impresora = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            kpi_impresora_red = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_impresora_red = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            # 4. OS Versions
            kpi_win7 = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND os_name LIKE '%Windows 7%'"
            ).fetchone()["c"]
            
            kpi_win10 = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND (os_name LIKE '%Windows 10%' OR os_name LIKE '%Windows 11%')"
            ).fetchone()["c"]

            # 5. Tareas hechas hoy
            kpi_tareas_hoy = conn.execute(
                "SELECT COUNT(*) as c FROM tasks WHERE estado = 'Hecha' AND DATE(completed_at) = DATE('now', 'localtime')"
            ).fetchone()["c"]
            
            # Tareas pendientes (para el card)
            # Nota: Esto es un calculo un poco mas pesado si hay muchas PCs, pero necesario para el KPI
            # Alternativa: count distinct pc_name from tasks where estado != hecho
            # Pero el template original usa: pcs|selectattr('tareas_pendientes', 'gt', 0)|list|length
            # Vamos a hacer una query directa más eficiente:
            kpi_tareas_pendientes_count = conn.execute(
                 "SELECT COUNT(DISTINCT pc_name) as c FROM tasks WHERE estado != 'Hecha' AND pc_name IS NOT NULL AND pc_name != ''"
            ).fetchone()["c"]

            # KPI Tareas Pendientes TOTAL (Asignadas + Sin Asignar)
            kpi_tareas_pendientes_total = conn.execute(
                "SELECT COUNT(*) as c FROM tasks WHERE estado != 'Hecha'"
            ).fetchone()["c"]


            # 6. Tareas por Categoria (para Charts)
            rows_cats = conn.execute("SELECT categoria, COUNT(*) as c FROM tasks GROUP BY categoria").fetchall()
            cat_labels = []
            cat_values = []
            for r in rows_cats:
                cat_name = r["categoria"] if r["categoria"] else "Sin Categoría"
                if cat_name == "General":
                    continue
                cat_labels.append(cat_name)
                cat_values.append(r["c"])

            # 7. Distribución SO (para Chart)
            # Reutilizamos kpi_win7 y win10, o hacemos query completa si hay otros SO
            # El original usa un canvas "osChart". Necesitamos pasar datos si queremos recrearlo.
            # Vamos a asumir que el JS pedía los datos? O se inyectaban?
            # En index.html original: <canvas id="osChart"></canvas>.
            # No veo donde se inyectaban los datos de charts en el python original en el bloque reemplazado...
            # Ah, espera, en render_template no se pasaban listas para charts excepto cat_labels/values?
            # Revisando el bloque eliminado: solo cat_labels/values.
            # El osChart quizas se llena via API o con variables JS inline que no vi?
            # Revisaré index.html para ver como se llenan los charts.
            
    except Exception as e:
        print(f"Error en graficos: {e}")
        kpi_total_activas = 0
        kpi_total_graveyard = 0
        kpi_alerta_ram = 0
        kpi_sin_impresora = 0
        kpi_impresora_red = 0
        kpi_win7 = 0
        kpi_win10 = 0
        kpi_tareas_hoy = 0
        kpi_tareas_pendientes_count = 0
        kpi_tareas_pendientes_total = 0
        cat_labels = []
        cat_values = []

    return render_template(
        "graficos.html",
        kpi_total_activas=kpi_total_activas,
        kpi_total_graveyard=kpi_total_graveyard,
        kpi_alerta_ram=kpi_alerta_ram,
        kpi_sin_impresora=kpi_sin_impresora,
        kpi_impresora_red=kpi_impresora_red,
        kpi_win7=kpi_win7,
        kpi_win10=kpi_win10,
        kpi_tareas_hoy=kpi_tareas_hoy,
        kpi_tareas_pendientes_count=kpi_tareas_pendientes_count,
        kpi_tareas_pendientes_total=kpi_tareas_pendientes_total,
        cat_labels=cat_labels,
        cat_values=cat_values,
        hostname=socket.gethostname(),
        fueros=FUERO_MAPPING  # Pasar mapping a template graficos/dashboard
    )

@app.route("/", methods=["GET"])
def dashboard():
    """Lista todas las PCs (activas y en cementerio) + KPIs + filtros + paginado."""
    pcs_data = []
    # KPIs se movieron a /graficos, inicializamos a 0 por si jinja los pide (aunque los quitaremos del template)
    kpi_tareas_hoy = 0
    kpi_tareas_pendientes_total = 0


    # Filtros
    q = request.args.get("q", "").strip()
    estado = request.args.get("estado", "True").strip()
    alerta = request.args.get("alerta", "").strip()
    os_param = request.args.get("os", "").strip()
    filter_tasks = request.args.get("filter_tasks", "").strip()
    
    # Ordenamiento
    sort_by = request.args.get("sort_by", "pc_name").strip()
    order = request.args.get("order", "asc").strip()

    # Paginación
    try:
        page = int(request.args.get("page", 1))
    except ValueError:
        page = 1
    
    try:
        per_page = int(request.args.get("per_page", 25))
    except ValueError:
        per_page = 25
        
    offset = (page - 1) * per_page

    total_rows = 0

    try:
        with get_db_connection() as conn:
            # Primero contamos cuántas filas habría con los filtros
            count_sql = "SELECT COUNT(*) as c FROM pcs p WHERE 1=1"
            params = []

            if q:
                count_sql += " AND p.pc_name LIKE ?"
                params.append(f"%{q}%")

            if estado in ("True", "False"):
                count_sql += " AND p.is_active = ?"
                params.append(estado)

            if alerta == "ram":
                count_sql += " AND p.alerta_ram_baja = 1"
            elif alerta == "sinimp":
                count_sql += " AND p.alerta_sin_impresora = 1"
            elif alerta == "red":
                count_sql += " AND p.alerta_impresora_red = 1"
            
            if os_param == "win7":
                 count_sql += " AND p.os_name LIKE '%Windows 7%'"
            elif os_param == "win10":
                 count_sql += " AND (p.os_name LIKE '%Windows 10%' OR p.os_name LIKE '%Windows 11%')"

            if filter_tasks == "true":
                count_sql += " AND (SELECT COUNT(*) FROM tasks t WHERE t.pc_name = p.pc_name AND t.estado != 'Hecha') > 0"

            total_rows = conn.execute(count_sql, params).fetchone()["c"]

            # Contar tareas pendientes sin asignar (para mostrar en dashboard)
            unassigned_tasks = conn.execute(
                "SELECT * FROM tasks WHERE pc_name IS NULL OR pc_name = '' AND estado != 'Hecha' ORDER BY created_at DESC"
            ).fetchall()
            unassigned_count = len(unassigned_tasks)
            
            # Obtener lista de técnicos
            technicians_list = [dict(r) for r in conn.execute("SELECT * FROM technicians ORDER BY name").fetchall()]

            # Ahora traemos solo la página actual
            base_sql = """
                SELECT
                    p.*,
                    (
                        SELECT COUNT(*)
                        FROM tasks t
                        WHERE t.pc_name = p.pc_name
                          AND t.estado != 'Hecha'
                    ) AS tareas_pendientes
                FROM pcs p
                WHERE 1=1
            """

            if q:
                base_sql += " AND p.pc_name LIKE ?"
            if estado in ("True", "False"):
                base_sql += " AND p.is_active = ?"
            if alerta == "ram":
                base_sql += " AND p.alerta_ram_baja = 1"
            elif alerta == "sinimp":
                base_sql += " AND p.alerta_sin_impresora = 1"
            elif alerta == "red":
                base_sql += " AND p.alerta_impresora_red = 1"
            
            if os_param == "win7":
                 base_sql += " AND p.os_name LIKE '%Windows 7%'"
            elif os_param == "win10":
                 base_sql += " AND (p.os_name LIKE '%Windows 10%' OR p.os_name LIKE '%Windows 11%')"
            
            if filter_tasks == "true":
                base_sql += " AND (SELECT COUNT(*) FROM tasks t WHERE t.pc_name = p.pc_name AND t.estado != 'Hecha') > 0"
            
            # Ordenamiento seguro (solo columnas permitidas)
            allowed_sort_cols = {
                "pc_name": "p.pc_name",
                "last_user": "p.last_user",
                "fuero": "p.fuero",
                "motherboard_model": "p.motherboard_model",
                "os_name": "p.os_name",
                "processor": "p.processor",
                "ram_gb": "p.ram_gb",
                "ram_detalles": "p.ram_detalles",
                "disk_models": "p.disk_models",
                "printer_model": "p.printer_model",
                "monitors": "p.monitors",
                "ip_address": "p.ip_address"
            }
            
            sort_col_sql = allowed_sort_cols.get(sort_by, "p.pc_name")
            sort_dir_sql = "DESC" if order == "desc" else "ASC"
            
            # Usar NULLS LAST para que los vacíos queden al final siempre
            base_sql += f" ORDER BY {sort_col_sql} {sort_dir_sql} NULLS LAST"

            # Añadir límite y offset para paginación
            base_sql += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])

            pcs_data = [dict(row) for row in conn.execute(base_sql, params).fetchall()]

            # --- RESTAURADO: CALCULO GLOBAL DE KPIS PARA DASHBOARD ---
            # 1. Total Activas (Excluyendo PC Generica e Infraestructura)
            kpi_total_activas = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND pc_name NOT IN ('PC Generica', 'Infraestructura')"
            ).fetchone()["c"]

            # 2. Total Cementerio
            kpi_total_graveyard = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'False'"
            ).fetchone()["c"]

            # 3. Alertas (RAM, Sin Impresora, Impresora Red)
            kpi_alerta_ram = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_ram_baja = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            kpi_sin_impresora = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_sin_impresora = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            kpi_impresora_red = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND alerta_impresora_red = 1 AND pc_name != 'PC Generica'"
            ).fetchone()["c"]

            # 4. OS Versions
            kpi_win7 = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND os_name LIKE '%Windows 7%'"
            ).fetchone()["c"]
            
            kpi_win10 = conn.execute(
                "SELECT COUNT(*) as c FROM pcs WHERE is_active = 'True' AND (os_name LIKE '%Windows 10%' OR os_name LIKE '%Windows 11%')"
            ).fetchone()["c"]

            # 5. Tareas hechas hoy
            kpi_tareas_hoy = conn.execute(
                "SELECT COUNT(*) as c FROM tasks WHERE estado = 'Hecha' AND DATE(completed_at) = DATE('now', 'localtime')"
            ).fetchone()["c"]

            # KPI Tareas Pendientes TOTAL (Asignadas + Sin Asignar)
            kpi_tareas_pendientes_total = conn.execute(
                "SELECT COUNT(*) as c FROM tasks WHERE estado != 'Hecha'"
            ).fetchone()["c"]
            
            # (Opcional) Si 'localtime' da problemas en producción/docker, usar DATE('now') o gestionar zona horaria en python.
            # Como corre en local Windows del usuario, 'localtime' debería tomar la hora del sistema.
            
    except Exception as exc:
        print(f"Error cargando dashboard: {exc}")
        technicians_list = []
        unassigned_tasks = []
        unassigned_count = 0
        kpi_tareas_pendientes_total = 0
        kpi_total_activas = 0
        kpi_total_graveyard = 0
        kpi_alerta_ram = 0
        kpi_sin_impresora = 0
        kpi_impresora_red = 0
        kpi_win7 = 0
        kpi_win10 = 0
        kpi_tareas_hoy = 0

    # Calcular total de páginas
    total_pages = (total_rows // per_page) + (1 if total_rows % per_page else 0)

    server_url = request.url_root.strip("/")  # ej: "http://192.168.1.8:5000"

    return render_template(
        "index.html",
        pcs=pcs_data,
        server_url=request.host_url,
        unassigned_tasks=unassigned_tasks,
        unassigned_count=unassigned_count,
        technicians=technicians_list,
        kpi_tareas_hoy=kpi_tareas_hoy,
        kpi_tareas_pendientes_total=kpi_tareas_pendientes_total,
        fueros=FUERO_MAPPING,  # Pasar mapping para filtros o modales
        # KPIs faltantes
        kpi_total_activas=kpi_total_activas,
        kpi_total_graveyard=kpi_total_graveyard,
        kpi_alerta_ram=kpi_alerta_ram,
        kpi_sin_impresora=kpi_sin_impresora,
        kpi_impresora_red=kpi_impresora_red,
        kpi_win7=kpi_win7,
        kpi_win10=kpi_win10,

        # Params de filtro para mantener estado en UI si se desea
        q=q,
        estado=estado,
        alerta=alerta,
        page=page,
        total_pages=total_pages,
        fuero_colors=FUERO_COLORS,
        per_page=per_page
    )


# --------- Exportar a CSV ---------

@app.route("/export", methods=["GET", "POST"])
def export_inventory():
    """GET: muestra formulario. POST: genera Excel con campos seleccionados."""
    
    if request.method == "GET":
        # Mostrar formulario para elegir campos
        return render_template("export_inventory.html")
    
    # POST: generar Excel
    campos_seleccionados = request.form.getlist("campos")
    
    if not campos_seleccionados:
        # Si no eligió nada, incluir los básicos
        campos_seleccionados = ["pc_name", "os_name", "processor", "ram_gb", "ip_address"]
    
    with get_db_connection() as conn:
        rows = conn.execute("SELECT * FROM pcs ORDER BY last_report DESC").fetchall()
    
    if not rows:
        return "Sin datos", 404
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Encabezados (solo los campos seleccionados)
    ws.append(campos_seleccionados)
    
    # Filas (solo los campos seleccionados)
    for row in rows:
        fila = [row[campo] for campo in campos_seleccionados]
        ws.append(fila)
    
    # Generar archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Inventario_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )



# --------- Detalle + Tareas ---------

@app.route("/pc/<pc_name>")
def pc_detail(pc_name):
    with get_db_connection() as conn:
        pc = conn.execute(
            "SELECT * FROM pcs WHERE pc_name = ?",
            (pc_name,),
        ).fetchone()

        tareas = conn.execute(
            """
            SELECT id, pc_name, created_at, descripcion, estado, solicitante, assigned_to
            FROM tasks
            WHERE pc_name = ?
            ORDER BY created_at DESC
            """,
            (pc_name,),
        ).fetchall()

        technicians = [dict(row) for row in conn.execute("SELECT * FROM technicians ORDER BY name").fetchall()]

        # Obtener historial de cambios
        audit_logs = conn.execute(
            "SELECT * FROM audit_logs WHERE pc_name = ? ORDER BY changed_at DESC",
            (pc_name,)
        ).fetchall()

        # Para el modal de migración (solo si es PC Generica, pero lo pasamos siempre por simplicidad o filtrado en jinja)
        all_pcs = conn.execute("SELECT pc_name, fuero, last_user FROM pcs WHERE is_active='True' ORDER BY pc_name").fetchall()

    if pc is None:
        abort(404)

    return render_template("pc_detail.html", pc=pc, tareas=tareas, technicians=technicians, audit_logs=audit_logs, all_pcs=all_pcs, fuero_colors=FUERO_COLORS)


@app.route("/pc/<pc_name>/update_infrastructure", methods=["POST"])
def update_pc_infrastructure(pc_name):
    """Actualiza los datos de infraestructura de una PC (Edificio, Piso, Switch, Pachera)."""
    building = request.form.get("building", "").strip()
    floor = request.form.get("floor", "").strip()
    switch_name = request.form.get("switch_name", "").strip()
    switch_port = request.form.get("switch_port", "").strip()
    pachera_name = request.form.get("pachera_name", "").strip()
    pachera_port = request.form.get("pachera_port", "").strip()
    
    try:
        with get_db_connection() as conn:
            # Verificar si existía para audit logs (opcional, pero recomendado)
            old_pc = conn.execute("SELECT * FROM pcs WHERE pc_name = ?", (pc_name,)).fetchone()
            
            conn.execute(
                """
                UPDATE pcs 
                SET building = ?, floor = ?, switch_name = ?, switch_port = ?, pachera_name = ?, pachera_port = ?
                WHERE pc_name = ?
                """,
                (building, floor, switch_name, switch_port, pachera_name, pachera_port, pc_name)
            )
            
            # Registrar cambios en audit_logs
            if old_pc:
                changes = [
                    ("building", old_pc["building"], building),
                    ("floor", old_pc["floor"], floor),
                    ("switch_name", old_pc["switch_name"], switch_name),
                    ("switch_port", old_pc["switch_port"], switch_port),
                    ("pachera_name", old_pc["pachera_name"], pachera_name),
                    ("pachera_port", old_pc["pachera_port"], pachera_port)
                ]
                
                for field, old, new in changes:
                    # Normalizar None a "" para comparación
                    old_str = str(old) if old is not None else ""
                    new_str = str(new) if new is not None else ""
                    
                    if old_str != new_str:
                        conn.execute(
                            "INSERT INTO audit_logs (pc_name, field, old_value, new_value) VALUES (?, ?, ?, ?)",
                            (pc_name, field, old_str, new_str)
                        )
            
            conn.commit()
            
        return redirect(url_for("pc_detail", pc_name=pc_name))
        
    except Exception as e:
        return f"Error actualizando infraestructura: {e}", 500


@app.route("/pc/migrate_tasks", methods=["POST"])
def migrate_generic_tasks():
    target_pc = request.form.get("target_pc")
    task_id = request.form.get("migration_task_id")

    if not target_pc:
        return redirect(url_for("pc_detail", pc_name="PC Generica"))
    
    with get_db_connection() as conn:
        if task_id:
            # Migrar una sola tarea
            conn.execute(
                "UPDATE tasks SET pc_name = ? WHERE id = ? AND pc_name = 'PC Generica'",
                (target_pc, task_id)
            )
            audit_msg = f"Se importó la tarea #{task_id} de PC Generica"
        else:
            # Fallback (o legado): Mover todas (aunque la UI ya no tenga el botón global)
            conn.execute(
                "UPDATE tasks SET pc_name = ? WHERE pc_name = 'PC Generica'",
                (target_pc,)
            )
            audit_msg = "Se importaron TODAS las tareas de PC Generica"

        # Registro en audit logs de la PC destino
        conn.execute(
            "INSERT INTO audit_logs (pc_name, field, old_value, new_value) VALUES (?, ?, ?, ?)",
            (target_pc, "MIGRACION", "PC Generica", audit_msg)
        )
        conn.commit()

    return redirect(url_for("pc_detail", pc_name=target_pc))


@app.route("/pc/<pc_name>/tasks", methods=["POST"])
def add_task(pc_name):
    descripcion = request.form.get("descripcion", "").strip()
    solicitante = request.form.get("solicitante", "").strip()
    categoria = request.form.get("categoria", "").strip()
    
    print(f"DEBUG: add_task called. Desc: {descripcion}, Solicitante Input: '{solicitante}'")

    # Enforce Solicitante
    if not solicitante:
        solicitante = "No Especificado (Dashboard)"

    if not descripcion:
        return redirect(url_for("pc_detail", pc_name=pc_name))

    if not categoria:
        categoria = predict_category(descripcion)

    with get_db_connection() as conn:
        conn.execute(
            "INSERT INTO tasks (pc_name, descripcion, solicitante, categoria) VALUES (?, ?, ?, ?)",
            (pc_name, descripcion, solicitante, categoria),
        )
        conn.commit()

    return redirect(url_for("pc_detail", pc_name=pc_name))


@app.route("/technicians/add", methods=["POST"])
def add_technician():
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db_connection() as conn:
                conn.execute("INSERT INTO technicians (name) VALUES (?)", (name,))
                conn.commit()
        except sqlite3.IntegrityError:
            pass # Ya existe
    return redirect(url_for("dashboard"))

@app.route("/technicians/delete/<int:tech_id>", methods=["POST"])
def delete_technician(tech_id):
    with get_db_connection() as conn:
        conn.execute("DELETE FROM technicians WHERE id = ?", (tech_id,))
        conn.commit()
    return redirect(url_for("dashboard"))


@app.route("/tasks/<int:task_id>/done", methods=["POST"])
def mark_task_done(task_id):
    technician = request.form.get("technician_name", None)
    
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT pc_name FROM tasks WHERE id = ?",
            (task_id,),
        ).fetchone()

        if row:
            conn.execute(
                """
                UPDATE tasks 
                SET estado = 'Hecha',
                    completed_by = ?,
                    completed_at = ?
                WHERE id = ?
                """,
                (technician, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), task_id),
            )
            conn.commit()
            pc_name = row["pc_name"]
        else:
            pc_name = ""

    if not pc_name:
        return redirect(url_for("dashboard"))

    return redirect(url_for("pc_detail", pc_name=pc_name))


@app.route("/tasks/<int:task_id>/delete", methods=["POST"])
def delete_task(task_id):
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT pc_name FROM tasks WHERE id = ?",
            (task_id,),
        ).fetchone()

        if row:
            conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
            conn.commit()
            pc_name = row["pc_name"]
        else:
            pc_name = ""

    if not pc_name:
        return redirect(url_for("dashboard"))

    return redirect(url_for("pc_detail", pc_name=pc_name))


@app.route("/create_loose_task", methods=["POST"])
def create_loose_task():
    """Crea una tarea sin PC asignada (global), pero ahora con posible FUERO."""
    descripcion = request.form.get("descripcion")
    solicitante = request.form.get("solicitante")
    categoria = request.form.get("categoria")
    technician = request.form.get("technician")
    fuero = request.form.get("fuero") # Nuevo campo

    if not descripcion or not solicitante:
        return "Faltan datos", 400
    
    # Si no elige categoría, intentar predecir con IA
    if not categoria:
        categoria = predict_category(descripcion)
    
    # Si elige técnico, el estado pasa a "Asignada" (aunque no tenga PC, tiene responsable)
    # SI no tiene PC ni técnico, es "Pendiente" (bolsa general)
    estado = "Pendiente"
    assigned_to = None
    
    if technician:
        assigned_to = technician
        estado = "Asignada"  # Ojo: Asignada a Técnico, pero sin PC.

    with get_db_connection() as conn:
        conn.execute(
            """INSERT INTO tasks (descripcion, solicitante, estado, created_at, categoria, assigned_to, fuero, pc_name)
               VALUES (?, ?, ?, datetime('now', 'localtime'), ?, ?, ?, 'PC Generica')""",
            (descripcion, solicitante, estado, categoria, assigned_to, fuero),
        )
        conn.commit()

    return redirect(url_for("dashboard"))


@app.route("/tasks/assign", methods=["POST"])
def assign_task():
    task_id = request.form.get("task_id")
    pc_name = request.form.get("pc_name", "").strip()
    
    if task_id and pc_name:
        with get_db_connection() as conn:
            # Verificar si existe la PC (opcional, pero buena práctica)
            t = conn.execute("SELECT 1 FROM pcs WHERE pc_name = ?", (pc_name,)).fetchone()
            if t:
                conn.execute(
                    "UPDATE tasks SET pc_name = ? WHERE id = ?",
                    (pc_name, task_id)
                )
                conn.commit()
            else:
                # Podríamos manejar error, por ahora redirigimos igual
                pass
                
    return redirect(url_for("dashboard"))


# --------- Reporte tareas a Excel ---------

@app.route("/report/tasks_completed", methods=["GET", "POST"])
def report_tasks_completed():
    # Si viene ?pc=... en la URL, lo usamos como filtro opcional
    pc_name = request.args.get("pc", "").strip()

    if request.method == "GET":
        # Mostrar formulario para elegir fecha (y mostrar si hay filtro de PC)
        return render_template("report_tasks.html", pc_name=pc_name)

    # POST: generar y descargar el Excel
    fecha_filtro = None
    fecha_str = request.form.get("fecha", "").strip()
    if fecha_str:
        fecha_filtro = fecha_str  # formato YYYY-MM-DD

    if not fecha_filtro:
        fecha_filtro = dt.now().strftime("%Y-%m-%d")

    # También permitimos que el formulario mande pc_name oculto
    pc_name_form = request.form.get("pc_name", "").strip()
    if pc_name_form:
        pc_name = pc_name_form

    with get_db_connection() as conn:
        base_sql = """
            SELECT pc_name, descripcion, solicitante, created_at, estado, completed_by, completed_at
            FROM tasks
            WHERE estado = 'Hecha'
              AND DATE(completed_at) = ?
        """
        params = [fecha_filtro]

        if pc_name:
            base_sql += " AND pc_name = ?"
            params.append(pc_name)

        base_sql += " ORDER BY completed_at DESC"

        tareas = conn.execute(base_sql, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"

    ws.append(["PC", "Descripción", "Solicitante", "Fecha Creación", "Estado", "Realizado Por", "Fecha Cierre"])

    for t in tareas:
        ws.append([
            t["pc_name"],
            t["descripcion"],
            t["solicitante"] or "",
            t["created_at"],
            t["estado"],
            t["completed_by"] or "",
            t["completed_at"] or ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_pc_sufijo = f"_{pc_name}" if pc_name else ""
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Tareas_Completadas{nombre_pc_sufijo}_{fecha_filtro}.xlsx",
    )


# --------- PDF Report Class ---------
from fpdf import FPDF

# Helper para fechas en español (sin depender de locale del sistema)
def format_date_es(d_obj):
    if isinstance(d_obj, str):
        try:
            d_obj = datetime.datetime.strptime(d_obj, "%Y-%m-%d")
        except:
            return d_obj
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    # Formato: Lunes 14 de Diciembre de 2025 ?? O User pidió "Lunes 14/12/2025"
    # "formato con el nombre de l dia mejor" -> Lunes 14/12/2025 is concise and good.
    return f"{dias[d_obj.weekday()]} {d_obj.day:02d}/{d_obj.month:02d}/{d_obj.year}"

def format_datetime_es(datetime_str):
    """Convierte datetime string a formato español DD/MM/YYYY HH:MM"""
    if not datetime_str:
        return ""
    try:
        # Parsear formato: "2025-12-30 14:30:00"
        dt_obj = datetime.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
        return f"{dt_obj.day:02d}/{dt_obj.month:02d}/{dt_obj.year} {dt_obj.hour:02d}:{dt_obj.minute:02d}"
    except:
        # Si falla, intentar extraer solo la hora como fallback
        if ' ' in datetime_str:
            return datetime_str.split(' ')[1][:5]
        return datetime_str

class PDFReport(FPDF):
    def __init__(self, title="Reporte - Inventario GOLD", orientation='P', unit='mm', format='A4'):
        super().__init__(orientation, unit, format)
        self.report_title = title

    def header(self):
        # Logo placeholder or Title
        self.set_font('Arial', 'B', 15)
        self.set_text_color(33, 37, 41) # Dark gray
        self.cell(0, 10, self.report_title, 0, 1, 'C')
        self.ln(5)
        
        # Line break
        self.set_draw_color(13, 110, 253) # Bootstrap Primary Blue
        self.set_line_width(1)
        current_width = self.w - 20 # margin 10 left right
        self.line(10, 25, 10 + current_width, 25)
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128)
        # Fecha simple en footer
        ahora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}} - Generado el {ahora}', 0, 0, 'C')

@app.route("/report/tasks_completed_pdf", methods=["POST"])
def report_tasks_completed_pdf():
    # ... query code ...
    fecha_filtro_str = request.form.get("fecha", "").strip() or dt.now().strftime("%Y-%m-%d")
    pc_name = request.form.get("pc_name", "").strip()
    
    # Parse date for display
    fecha_dt = datetime.datetime.strptime(fecha_filtro_str, "%Y-%m-%d")
    fecha_display = format_date_es(fecha_dt)

    with get_db_connection() as conn:
        # 1. Tareas REALIZADAS (Hecha)
        base_sql = """
            SELECT t.pc_name, t.descripcion, t.solicitante, t.created_at, t.estado, t.completed_by, t.completed_at,
                   p.last_user
            FROM tasks t
            LEFT JOIN pcs p ON t.pc_name = p.pc_name
            WHERE t.estado = 'Hecha'
              AND DATE(t.completed_at) = ?
        """
        params = [fecha_filtro_str]

        if pc_name:
            base_sql += " AND t.pc_name = ?"
            params.append(pc_name)

        base_sql += " ORDER BY t.completed_at DESC"
        tareas_hechas = conn.execute(base_sql, params).fetchall()

        # 2. Tareas PENDIENTES
        sql_pendientes = """
            SELECT t.pc_name, t.descripcion, t.solicitante, t.created_at, t.estado, t.assigned_to,
                   p.last_user
            FROM tasks t
            LEFT JOIN pcs p ON t.pc_name = p.pc_name
            WHERE t.estado != 'Hecha'
              AND DATE(t.created_at) = ?
        """
        params_pend = [fecha_filtro_str]
        if pc_name:
            sql_pendientes += " AND t.pc_name = ?"
            params_pend.append(pc_name)
        
        sql_pendientes += " ORDER BY t.created_at DESC"
        tareas_pendientes = conn.execute(sql_pendientes, params_pend).fetchall()

    # --- Generación PDF ---
    pdf = PDFReport(title="Reporte de Tareas - Inventario GOLD")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Titulo del día
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Reporte del día: {fecha_display}", 0, 1)
    pdf.ln(2)

    # --- SECCIÓN 1: TAREAS REALIZADAS ---
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(25, 135, 84) # Success Green
    pdf.cell(0, 8, f"Tareas Realizadas ({len(tareas_hechas)})", 0, 1)
    pdf.ln(2)

    # Headers
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(25, 135, 84) # Green header
    pdf.set_text_color(255)
    
    # PC(28), Usuario(22), Desc(45), Solic(18), Tech(18), Fecha Creada(34)
    headers = ["PC", "Usuario", "Descripción", "Solic.", "Técnico", "Fecha Creada"]
    w = [28, 22, 45, 18, 18, 34]
    
    for i, h in enumerate(headers):
        pdf.cell(w[i], 8, h, 1, 0, 'C', fill=True)
    pdf.ln()

    if not tareas_hechas:
        pdf.set_font("Arial", "I", 9)
        pdf.set_text_color(0)
        pdf.cell(0, 8, "No hay tareas realizadas registradas para hoy.", 1, 1, 'C')
    else:
        pdf.set_font("Arial", "", 8)
        pdf.set_text_color(0)
        for t in tareas_hechas:
            # Procesar datos
            raw_user = t["last_user"] or "N/A"
            if "\\" in raw_user: user_display = raw_user.split("\\")[-1]
            else: user_display = raw_user
            
            desc = t["descripcion"] or ""
            solicitante = t["solicitante"] or ""
            tecnico = t["completed_by"] or ""
            
            # Fecha
            created_at = t["created_at"] or ""
            fecha_hora = format_datetime_es(created_at)
            
            pc_name = str(t["pc_name"])
            user_display = str(user_display)

            # CALCULAR ALTURA DE LA FILA (Maximo entre Descripcion y Solicitante)
            # Aproximacion simple: chars / ancho.
            # Mejor: Usar MultiCell en "dry run" o guardar Y.
            
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            
            # Simulamos altura renderizando la descripcion (la mas larga)
            # FPDF no tiene "dry run" facil sin hacks. 
            # Estrategia: Guardar Y, imprimir celda ficticia? No.
            # Estrategia: Contar lineas basado en width aprox.
            # Width descripcion: 45. Width solicitante: 18.
            # Font size 8 aprox 2mm width avg char? -> NO, variable.
            # Usaremos GetStringWidth si es posible, o simplemente dejaremos que FPDF maneje el cursor.
            
            # MEJOR ESTRATEGIA: Renderizar las MultiCells primero, ver cual es la mas alta, y luego las celdas simples.
            # Pero las celdas simples deben tener esa altura.
            
            # 1. Calcular altura necesaria
            # Contamos caracteres y dividimos por un promedio seguro (e.g. 2.5mm per char? Aprox 18 chars en 45mm con size 8)
            # Desc limit width 45. Solicitante limit width 18.
            
            # Determinar altura maxima de la fila
            # Nota: FPDF mueve la pagina si hay salto. Esto complica las cosas si estamos al final de pagina.
            # Asumiremos que add_page automatico funciona, pero el "y_start" cambiaria.
            # Para simplificar: usaremos height fijo calculado por len() para evitar complejidad de salto de pagina manual.
            
            # Revertimos a estrategia mas segura para este contexto limitado:
            # Calcular height basado en chars (heuristica).
            # Font size 8. width 45mm (~127pt). Avg char width ~3-4pt. ~30-40 chars por linea.
            lines_desc = max(1, (len(desc) // 25) + 1)
            lines_solic = max(1, (len(solicitante) // 9) + 1) # width 18 espoco, ~9 chars
            
            # Altura base 5mm por linea
            max_lines = max(lines_desc, lines_solic)
            h_row = max_lines * 5
            
            # Check page break
            if (y_start + h_row) > 275: # A4 height ~297 minus margins
                pdf.add_page()
                y_start = pdf.get_y()
                # Redraw headers? (Opcional, pero ideal)
            
            # Renderizar Celdas (Ahora si)
            pdf.set_xy(x_start, y_start)
            
            # PC
            pdf.cell(w[0], h_row, pc_name[:16], 1)
            
            # User
            pdf.cell(w[1], h_row, user_display[:13], 1)
            
            # Desc (MultiCell)
            x_desc = x_start + w[0] + w[1]
            pdf.set_xy(x_desc, y_start)
            pdf.multi_cell(w[2], 5, desc, 0, 'L')
            # Si multicell ocupo menos que h_row, el borde queda mal?
            # MultiCell dibuja caja alrededor de texto. Si queremos caja full height:
            # Dibujamos Rect border despues.
            
            # Solic (MultiCell)
            x_solic = x_desc + w[2]
            pdf.set_xy(x_solic, y_start)
            pdf.multi_cell(w[3], 5, solicitante, 0, 'L')
            
            # Tecnico
            x_tech = x_solic + w[3]
            pdf.set_xy(x_tech, y_start)
            pdf.cell(w[4], h_row, tecnico[:11], 1)
            
            # Fecha
            x_date = x_tech + w[4]
            pdf.set_xy(x_date, y_start)
            pdf.cell(w[5], h_row, fecha_hora, 1, 1, 'C') # El ultimo param 1 salto linea? No, forzamos set_xy abajo
            
            # Dibujar rectangulos vacios alrededor de MultiCells para completar el grid si quedo corto
            pdf.rect(x_desc, y_start, w[2], h_row)
            pdf.rect(x_solic, y_start, w[3], h_row)
            
            # Moverse a la siguiente linea logicamente
            pdf.set_xy(x_start, y_start + h_row)
    
    pdf.ln(2)

    # --- SECCIÓN 2: TAREAS PENDIENTES ---
    pdf.set_font("Arial", "B", 11)
    pdf.set_text_color(220, 53, 69) # Danger Red
    pdf.cell(0, 8, f"Tareas Pendientes / Generadas Hoy ({len(tareas_pendientes)})", 0, 1)
    pdf.ln(2)

    # Headers Pendientes (Sin PC ni Usuario)
    # Desc(80), Solic(40), Técnico(40), Hora(25) -> Total 185
    headers_pend = ["Descripción", "Solic.", "Asignado a", "Fecha Creada"]
    w_pend = [70, 30, 30, 35]

    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(220, 53, 69) # Red header
    pdf.set_text_color(255)
    
    for i, h in enumerate(headers_pend):
        pdf.cell(w_pend[i], 8, h, 1, 0, 'C', fill=True)
    pdf.ln()

    if not tareas_pendientes:
        pdf.set_font("Arial", "I", 9)
        pdf.set_text_color(0)
        pdf.cell(0, 8, "No hay tareas pendientes generadas hoy.", 1, 1, 'C')
    else:
        pdf.set_font("Arial", "", 8)
        pdf.set_text_color(0)
        for t in tareas_pendientes:
            desc = t["descripcion"] or ""
            solicitante = str(t["solicitante"] or "")
            assigned = str(t["assigned_to"] or "Sin Asignar")
            
            created_at = t["created_at"] or ""
            fecha_hora = format_datetime_es(created_at)

            # CALCULO ALTURA (misma logica heuristica)
            # Widths: Desc(70), Solic(30), Asignado(30), Fecha(35)
            # Desc(70): ~35 chars por linea
            # Solic(30): ~15 chars por linea
            
            lines_desc = max(1, (len(desc) // 35) + 1)
            lines_solic = max(1, (len(solicitante) // 15) + 1)
            
            max_lines = max(lines_desc, lines_solic)
            h_row = max_lines * 5
            
            # Check page break
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            
            if (y_start + h_row) > 275:
                pdf.add_page()
                y_start = pdf.get_y()
            
            # Desc (MultiCell)
            pdf.set_xy(x_start, y_start)
            pdf.multi_cell(w_pend[0], 5, desc, 0, 'L')
            
            # Solic (MultiCell)
            x_solic = x_start + w_pend[0]
            pdf.set_xy(x_solic, y_start)
            pdf.multi_cell(w_pend[1], 5, solicitante, 0, 'L')
            
            # Asignado
            x_assign = x_solic + w_pend[1]
            pdf.set_xy(x_assign, y_start)
            pdf.cell(w_pend[2], h_row, assigned[:18], 1)
            
            # Fecha
            x_date = x_assign + w_pend[2]
            pdf.set_xy(x_date, y_start)
            pdf.cell(w_pend[3], h_row, fecha_hora, 1, 1, 'C')
            
            # Rectangulos borde para MultiCells
            pdf.rect(x_start, y_start, w_pend[0], h_row)
            pdf.rect(x_solic, y_start, w_pend[1], h_row)
            
            # Next line
            pdf.set_xy(x_start, y_start + h_row)


    # Output
    output = BytesIO()
    # fpdf2 output to string/bytes requires 'dest=S'. 
    # Actually fpdf2 has output(dest='S').encode('latin-1') logic. 
    # For using BytesIO with flask send_file, we can do:
    pdf_bytes = pdf.output()  # retorna bytearray en fpdf2 recientes o string
    # check fpdf2 version. Assuming default behavior: returns bytes.
    
    # Write to BytesIO
    output.write(pdf_bytes)
    output.seek(0)
    
    nombre_pc_sufijo = f"_{pc_name}" if pc_name else ""
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"Reporte_Tareas{nombre_pc_sufijo}_{fecha_filtro_str}.pdf",
    )


@app.route("/export_inventory_pdf", methods=["POST"])
def export_inventory_pdf():
    # Ignoramos selección dinámica para PDF, usamos columnas fijas "Pro"
    # Columnas: PC Name, Last User, OS Name, Processor, RAM, IP
    with get_db_connection() as conn:
        rows = conn.execute("SELECT pc_name, last_user, os_name, processor, ram_gb, ip_address FROM pcs WHERE pc_name != 'PC Generica' ORDER BY last_report DESC").fetchall()
    
    pdf = PDFReport(title="Inventario General - Inventario GOLD", orientation='L')
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Fecha de reporte
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 8, f"Fecha de emisión: {format_date_es(datetime.datetime.now())}", 0, 1, 'C')
    pdf.ln(5)
    
    # Titulo (Conteo)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Total Equipos Inventariados: {len(rows)}", 0, 1, 'C')
    pdf.ln(5)
    
    # Headers
    # Landscape A4 ~ 297mm ancho. Margen 10+10=20. Util: 277.
    # PC(40), User(40), OS(50), CPU(70), RAM(20), IP(35) -> Total 255
    headers = ["Nombre PC", "Usuario", "Sistema Operativo", "Procesador", "RAM", "IP Address"]
    w = [40, 40, 50, 70, 20, 35]
    
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(13, 110, 253)
    pdf.set_text_color(255)
    
    for i, h in enumerate(headers):
        pdf.cell(w[i], 8, h, 1, 0, 'C', fill=True)
    pdf.ln()
    
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(0)
    
    for row in rows:
        # User clean
        raw_user = row["last_user"] or "N/A"
        if "\\" in raw_user: user = raw_user.split("\\")[-1]
        else: user = raw_user
        
        # CPU truncate
        cpu = row["processor"] or "N/A"
        if len(cpu) > 40: cpu = cpu[:37] + "..."
        
        pdf.cell(w[0], 7, str(row["pc_name"]), 1)
        pdf.cell(w[1], 7, str(user)[:20], 1)
        pdf.cell(w[2], 7, str(row["os_name"]), 1)
        pdf.cell(w[3], 7, cpu, 1)
        pdf.cell(w[4], 7, f'{row["ram_gb"]} GB', 1, 0, 'C')
        pdf.cell(w[5], 7, str(row["ip_address"]), 1, 1, 'C')

    output = BytesIO()
    pdf_bytes = pdf.output()
    output.write(pdf_bytes)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"Inventario_Reporte_{dt.now().strftime('%Y%m%d')}.pdf",
    )



# --------- Acciones de estado ---------

@app.route("/decommission/<string:pc_name>", methods=["GET"])
def decommission_pc(pc_name):
    with get_db_connection() as conn:
        conn.execute(
            "UPDATE pcs SET is_active = 'False' WHERE pc_name = ?", (pc_name,)
        )
        conn.commit()
    return redirect(url_for("dashboard"))


@app.route("/reactivate/<pc_name>")
def reactivate_pc(pc_name):
    """Reactivar una PC (sacarla del cementerio)."""
    try:
        with get_db_connection() as conn:
            conn.execute(
                "UPDATE pcs SET is_active = 'True' WHERE pc_name = ?",
                (pc_name,),
            )
            conn.commit()
    except Exception as exc:
        print(f"Error reactivating PC {pc_name}: {exc}")

    return redirect(url_for("dashboard"))


# ----------------- API: inventario -----------------


def process_inventory_data(data):
    """Lógica central para procesar el JSON de inventario e insertar/actualizar en BD."""
    pc_name = data.get("PC_Nombre")
    if not pc_name:
        raise ValueError("Falta PC_Nombre en el JSON")

    last_user = data.get("Usuario_Actual", "N/A")
    fecha_raw = data.get("Fecha_Reporte", str(datetime.datetime.now()))
    last_report = fecha_raw["value"] if isinstance(fecha_raw, dict) else str(fecha_raw)

    sistema = data.get("Sistema", {})
    os_name = sistema.get("OsName", "N/A")
    processor = sistema.get("Procesador", "N/A")
    ram_gb = sistema.get("RAM (GB)", 0)

    # Red
    red = data.get("Red", [])
    ip_address = red[0].get("IPAddress") if red else "N/A"

    # Otros detalles
    ram_detalles = data.get("RAM_Detalles", "N/A")
    disk_models = data.get("Disk_Models", "N/A")
    disk_speeds_rpm = data.get("Disk_Speeds_RPM", "N/A")
    motherboard_model = data.get("Motherboard_Model", "N/A")
    printer_model = data.get("Printer_Model", "N/A")
    printer_port = data.get("Printer_Port", "N/A")
    monitors = data.get("Monitors", "N/A")

    # Calidad de conexión (desactivada, dejamos N/A)
    ping_ms = "N/A"
    ping_loss_pct = "N/A"

    # ------------ ALERTAS ------------

    # RAM baja
    try:
        ram_val = float(ram_gb)
    except Exception:
        ram_val = 0.0
    alerta_ram_baja = 1 if ram_val < 8 else 0

    # Impresoras
    pm = (printer_model or "").upper()
    pp = (printer_port or "").upper()

    sin_modelo = pm in ("", "N/A", "NINGUNA")

    # Impresoras virtuales típicas
    es_virtual = ("PDF" in pm) or ("XPS" in pm) or ("ONENOTE" in pm)

    es_red = ("IP_" in pp) or ("WSD" in pp) or ("\\" in pp)
    es_local = pp.startswith("USB") or pp.startswith("LPT")
    
    # Detectar desconexión física (reportada por script PS updated)
    esta_desconectada = "DESCONECTADA" in pp

    # Sin impresora física propia (o desconectada):
    if sin_modelo or es_virtual or esta_desconectada:
        alerta_sin_impresora = 1
    else:
        alerta_sin_impresora = 0

    # Solo impresora en red (física)
    alerta_impresora_red = 1 if (not sin_modelo and not es_virtual and es_red and not es_local) else 0

    # JSON completo
    full_json = json.dumps(data, ensure_ascii=False)

    # Detectar Fuero automáticamente
    fuero_detectado = detect_fuero(pc_name)

    sql = """
    INSERT INTO pcs (
        pc_name,
        fuero,
        os_name,
        processor,
        ram_gb,
        ip_address,
        last_user,
        last_report,
        ram_detalles,
        disk_models,
        disk_speeds_rpm,
        motherboard_model,
        monitors,
        printer_model,
        printer_port,
        ping_ms,
        ping_loss_pct,
        alerta_ram_baja,
        alerta_sin_impresora,
        alerta_impresora_red,
        is_active,
        full_json_data
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'True', ?)
    ON CONFLICT(pc_name) DO UPDATE SET
        fuero = excluded.fuero,
        os_name = excluded.os_name,
        processor = excluded.processor,
        ram_gb = excluded.ram_gb,
        ip_address = excluded.ip_address,
        last_user = excluded.last_user,
        last_report = excluded.last_report,
        ram_detalles = excluded.ram_detalles,
        disk_models = excluded.disk_models,
        disk_speeds_rpm = excluded.disk_speeds_rpm,
        motherboard_model = excluded.motherboard_model,
        monitors = excluded.monitors,
        printer_model = excluded.printer_model,
        printer_port = excluded.printer_port,
        ping_ms = excluded.ping_ms,
        ping_loss_pct = excluded.ping_loss_pct,
        alerta_ram_baja = excluded.alerta_ram_baja,
        alerta_sin_impresora = excluded.alerta_sin_impresora,
        alerta_impresora_red = excluded.alerta_impresora_red,
        full_json_data = excluded.full_json_data
    """

    with get_db_connection() as conn:
        # 1. Detectar cambios antes del UPDATE
        current_pc = conn.execute("SELECT * FROM pcs WHERE pc_name = ?", (pc_name,)).fetchone()
        
        if current_pc:
            fields_to_check = {
                "ram_gb": str(current_pc["ram_gb"]),
                "disk_models": str(current_pc["disk_models"]),
                "processor": str(current_pc["processor"]),
                "os_name": str(current_pc["os_name"]),
                "ip_address": str(current_pc["ip_address"])
            }
            
            new_values_map = {
                "ram_gb": str(ram_gb),
                "disk_models": str(disk_models),
                "processor": str(processor),
                "os_name": str(os_name),
                "ip_address": str(ip_address)
            }

            for field, old_val in fields_to_check.items():
                new_val = new_values_map.get(field, "N/A")
                if old_val.strip() != new_val.strip():
                    print(f"CAMBIO DETECTADO en {pc_name}: {field} de '{old_val}' a '{new_val}'")
                    conn.execute(
                        "INSERT INTO audit_logs (pc_name, field, old_value, new_value, changed_at) VALUES (?, ?, ?, ?, datetime('now', '-3 hours'))",
                        (pc_name, field, old_val, new_val)
                    )

        # 2. Ejecutar el Upsert
        conn.execute(
            sql,
            (
                pc_name,
                fuero_detectado,
                os_name,
                processor,
                ram_gb,
                ip_address,
                last_user,
                last_report,
                ram_detalles,
                disk_models,
                disk_speeds_rpm,
                motherboard_model,
                monitors,
                printer_model,
                printer_port,
                ping_ms,
                ping_loss_pct,
                alerta_ram_baja,
                alerta_sin_impresora,
                alerta_impresora_red,
                full_json,
            ),
        )
        conn.commit()

    print(f"Inventario guardado / actualizado: {pc_name}")
    return pc_name


@app.route("/submit_inventory", methods=["POST"])
def receive_inventory():
    """Recibe JSON de inventario y hace upsert en pcs."""
    try:
        raw_data = request.get_data()
        try:
            data = json.loads(raw_data.decode("utf-8"))
        except Exception:
            data = json.loads(raw_data.decode("utf-16"))

        process_inventory_data(data)
        return jsonify({"status": "success"}), 200

    except ValueError as ve:
        return jsonify({"status": "error", "message": str(ve)}), 400
    except Exception as exc:
        print(f"Error procesando inventario: {exc}")
        return jsonify({"status": "error", "message": str(exc)}), 500


@app.route("/upload_manual", methods=["POST"])
def upload_manual_inventory():
    """Permite subir un archivo JSON manualmente (Modo Offline)."""
    if 'file' not in request.files:
        return redirect(url_for('dashboard'))
    
    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('dashboard'))

    if file:
        try:
            # Leer el archivo
            content = file.read()
            # Intentar decodificar (puede venir en utf-8 o utf-16 desde powershell)
            try:
                data = json.loads(content.decode("utf-8"))
            except:
                data = json.loads(content.decode("utf-16"))
            
            pc_name = process_inventory_data(data)
            # Podríamos añadir un flash message aquí si tuviéramos soporte, 
            # pero por ahora redirigimos.
            print(f"Manual upload success for {pc_name}")
        except Exception as e:
            print(f"Error en carga manual: {e}")
            
        return redirect(url_for('dashboard'))


@app.route("/script")
def get_script():
    """Devuelve el contenido del script inventario.ps1 para ser copiado."""
    try:
        # Asegurarse de que el mimetype sea text/plain para que fetch lo lea como texto
        return send_file("inventario.ps1", mimetype="text/plain", as_attachment=False)
            
    except Exception as e:
        return f"Error al leer script: {e}", 500




# ----------------- API: logs -----------------

@app.route("/submit_log", methods=["POST"])
def receive_log():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"status": "error", "message": "No JSON"}), 400

        pc_name = data.get("PC_Nombre", "Unknown")
        log_content_raw = data.get("Log_Content", "No Content")

        if isinstance(log_content_raw, (dict, list)):
            log_content = json.dumps(log_content_raw, indent=4, ensure_ascii=False)
        else:
            log_content = str(log_content_raw)

        filename = f"LOG_{pc_name}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(LOG_FOLDER, filename)

        with open(filepath, "w", encoding="utf-8") as fh:
            fh.write(log_content)

        print(f"Log de error guardado: {filename}")
        return jsonify({"status": "success"}), 200

    except Exception as exc:
        print(f"Error guardando log: {exc}")
        return jsonify({"status": "error", "details": str(exc)}), 200


# ----------------- MOBILE API -----------------

@app.route("/mobile")
def mobile_view():
    return render_template("mobile.html")

@app.route("/api/mobile/data")
def api_mobile_data():
    """Returns technicians, unassigned tasks, and pcs for mobile view."""
    try:
        with get_db_connection() as conn:
            # Technicians
            techs = [dict(r) for r in conn.execute("SELECT * FROM technicians ORDER BY name").fetchall()]
            
            # Tasks assigned to specific tech (filtered by frontend usually, or we can filter here if param provided)
            # Unassigned tasks
            unassigned = [dict(r) for r in conn.execute(
                "SELECT * FROM tasks WHERE (pc_name IS NULL OR pc_name = '') AND estado != 'Hecha' ORDER BY created_at DESC"
            ).fetchall()]

            # All active tasks (Joined with PCs to get info like fuero)
            all_active = [dict(r) for r in conn.execute(
                """
                SELECT t.*, p.fuero as pc_fuero 
                FROM tasks t
                LEFT JOIN pcs p ON t.pc_name = p.pc_name
                WHERE t.estado != 'Hecha' 
                ORDER BY t.created_at DESC
                """
            ).fetchall()]

            # PCs for dropdown
            pcs = [r["pc_name"] for r in conn.execute("SELECT pc_name FROM pcs WHERE is_active='True' ORDER BY pc_name").fetchall()]

        return jsonify({
            "technicians": techs,
            "unassigned": unassigned,
            "active_tasks": all_active,
            "pcs": pcs
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/mobile/create_task", methods=["POST"])
def api_mobile_create_task():
    try:
        data = request.json
        descripcion = data.get("descripcion")
        pc_name = data.get("pc_name")
        technician = data.get("technician")
        solicitante_input = data.get("solicitante", "").strip()
        is_done = data.get("is_done", False)
        
        # Nuevos campos para infraestructura
        es_infraestructura = data.get("es_infraestructura", False)
        tecnico_ejecutor = data.get("tecnico", "")

        if not descripcion or not technician:
            return jsonify({"status": "error", "message": "Faltan datos"}), 400

        # Si es tarea de infraestructura y no tiene PC asignada, usar "Infraestructura"
        if es_infraestructura and not pc_name:
            pc_name = "Infraestructura"

        # Predict category if valid
        categoria = predict_category(descripcion)

        estado = "Hecha" if is_done else "Pendiente"
        created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        completed_at = created_at if is_done else None
        
        # Si hay técnico ejecutor y está marcada como hecha, usar ese técnico
        if is_done and tecnico_ejecutor:
            completed_by = tecnico_ejecutor
        elif is_done:
            completed_by = technician
        else:
            completed_by = None
        
        # Logic for Solicitante
        if solicitante_input:
            solicitante = solicitante_input
        else:
            solicitante = "No Especificado (Móvil)"

        assigned_to = technician # Auto-assign new mobile task to the creator

        with get_db_connection() as conn:
             cursor = conn.execute(
                """
                INSERT INTO tasks (pc_name, descripcion, solicitante, estado, created_at, completed_by, completed_at, categoria, assigned_to)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (pc_name, descripcion, solicitante, estado, created_at, completed_by, completed_at, categoria, assigned_to)
            )
             new_id = cursor.lastrowid
             conn.commit()

        return jsonify({"status": "success", "id": new_id})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/mobile/update_task", methods=["POST"])
def api_mobile_update_task():
    try:
        data = request.json
        task_id = data.get("id")
        action = data.get("action") # 'claim' or 'complete'
        technician = data.get("technician")
        pc_name = data.get("pc_name") # Optional, for updating PC when completing

        if not task_id or not action or not technician:
             return jsonify({"status": "error", "message": "Datos incompletos"}), 400

        with get_db_connection() as conn:
            if action == "claim":
                conn.execute(
                    "UPDATE tasks SET assigned_to=? WHERE id=?",
                    (technician, task_id)
                )
            elif action == "complete":
                 # Prepare query args
                 sql = "UPDATE tasks SET estado='Hecha', completed_by=?, completed_at=datetime('now', 'localtime')"
                 params = [technician]

                 # If PC name provided, update it too
                 if pc_name:
                     sql += ", pc_name=?"
                     params.append(pc_name)
                 
                 sql += " WHERE id=?"
                 params.append(task_id)

                 conn.execute(sql, params)
            
            conn.commit()
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500




# ----------------- Rutas de Instalación / Descarga -----------------

@app.route("/install")
def install_page():
    """Página simple para descargar los scripts del cliente."""
    host = request.host # Mantiene el puerto
    return f"""
    <html>
    <head>
        <title>Instalar Inventario</title>
        <style>
            body {{ font-family: sans-serif; padding: 40px; max-width: 600px; margin: 0 auto; background: #f8f9fa; }}
            .card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
            h1 {{ color: #0d6efd; margin-top: 0; }}
            a.btn {{ display: block; background: #0d6efd; color: white; padding: 15px; text-decoration: none; border-radius: 5px; margin: 10px 0; text-align: center; font-weight: bold; }}
            a.btn:hover {{ background: #0b5ed7; }}
            .step {{ margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="card">
            <h1>📥 Instalación Cliente</h1>
            <p>Sigue estos pasos en la PC que quieres inventariar (Windows 7/10/11):</p>
            
            <div class="step">
                <strong>1. Crea una carpeta</strong> en el Escritorio llamada <code>Inventario</code>.
            </div>

            <div class="step">
                <strong>2. Descarga los archivos</strong> en esa carpeta:
                <a href="/download/script" class="btn">📄 1. Descargar Script (inventario.ps1)</a>
                <a href="/download/launcher" class="btn">🚀 2. Descargar Ejecutable (ejecutar_inventario.bat)</a>
            </div>

            <div class="step">
                <strong>3. Ejecuta</strong> el archivo <code>ejecutar_inventario.bat</code> (doble clic).
            </div>
            
            <hr>
            <p><small>Si Windows protege la PC, pulsa "Más información" -> "Ejecutar de todas formas".</small></p>
        </div>
    </body>
    </html>
    """

@app.route("/download/script")
def download_client_script():
    try:
        # Leemos el archivo original
        with open("inventario.ps1", "r", encoding="utf-8") as f:
            content = f.read()
            
        # Reemplazamos 'localhost:5000' por la dirección actual (host:puerto)
        # Esto permite que si entras por 192.168.1.X, el script se baje con esa IP.
        current_host = request.host
        modified_content = content.replace("localhost:5000", current_host)
        
        # Servimos el contenido modificado desde memoria
        mem = BytesIO()
        mem.write(modified_content.encode("utf-8"))
        mem.seek(0)
        
        return send_file(mem, as_attachment=True, download_name="inventario.ps1")
    except Exception as e:
        return f"Error: {e}", 404

@app.route("/download/launcher")
def download_client_launcher():
    try:
        return send_file("ejecutar_inventario.bat", as_attachment=True, download_name="ejecutar_inventario.bat")
    except Exception as e:
        return f"Error: {e}", 404


@app.route("/download-cert")
def download_certificate():
    """Permite descargar el certificado SSL para instalarlo en dispositivos móviles."""
    try:
        return send_file(
            "cert.pem",
            as_attachment=True,
            download_name="inventario-cert.crt",
            mimetype="application/x-x509-ca-cert"
        )
    except Exception as e:
        return f"Error: {e}", 404



if __name__ == "__main__":
    import platform
    import threading
    sistema = platform.system()
    
    if sistema == "Windows":
        print("\n" + "="*64)
        print(" MODO DESARROLLO (Windows)")
        print(" Iniciando servidor Flask (Debug On)...")
        print("="*64)
        app.run(host="0.0.0.0", port=5000, debug=True)
    else:
        print("\n" + "="*64)
        print(" MODO PRODUCCIÓN (Linux)")
        print(" Iniciando servidor Flask con HTTPS y HTTP...")
        print(" - HTTPS: https://10.15.2.251:5000 (para PCs)")
        print(" - HTTP:  http://10.15.2.251:8080 (para móviles)")
        print("="*64)
        
        # Función para ejecutar servidor HTTPS
        def run_https():
            app.run(host="0.0.0.0", port=5000, debug=False, 
                    ssl_context=('cert.pem', 'key.pem'), use_reloader=False)
        
        # Función para ejecutar servidor HTTP
        def run_http():
            app.run(host="0.0.0.0", port=8080, debug=False, use_reloader=False)
        
        # Iniciar HTTPS en un thread separado
        https_thread = threading.Thread(target=run_https, daemon=True)
        https_thread.start()
        
        # Iniciar HTTP en el thread principal
        run_http()

